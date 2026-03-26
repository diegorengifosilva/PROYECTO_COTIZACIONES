
# ─── Librerías estándar ─────────────────────────────
import os
import io
import re
import json
import unicodedata
import pandas as pd
import platform
import subprocess
import logging
from decimal import Decimal, InvalidOperation
from html import unescape
from openpyxl import Workbook
from weasyprint import HTML
from pdf2docx import Converter
import tempfile
from pathlib import Path
import shutil
from docxtpl import DocxTemplate, RichText
import jinja2

# ─── Librerías de terceros ──────────────────────────
from reportlab.pdfgen import canvas

from . import serializers
logger = logging.getLogger(__name__)

# ─── Django core ────────────────────────────────────
from django.conf import settings
from django.http import JsonResponse, HttpResponse, FileResponse
from django.shortcuts import get_object_or_404
from django.views.decorators.csrf import ensure_csrf_cookie, csrf_exempt
from django.utils import timezone
from django.db.models import Sum, Count, Q, F, Max, DecimalField, ExpressionWrapper
from django.db.models.functions import TruncDate, Coalesce, ExtractMonth
from django.core.exceptions import ValidationError
from django.core.cache import cache
from django.db import transaction
from django.views.decorators.http import require_GET
from django.utils.dateparse import parse_date
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from rest_framework_simplejwt.authentication import JWTAuthentication

# ─── Django REST Framework ──────────────────────────
from rest_framework.decorators import api_view, parser_classes, permission_classes, action, authentication_classes
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework import status, viewsets, generics, filters, permissions
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.views import APIView
from rest_framework.exceptions import ValidationError
from rest_framework.pagination import PageNumberPagination
from rest_framework.generics import RetrieveAPIView

from django.http import HttpResponse
from django.template.loader import render_to_string
from django.conf import settings

from django.shortcuts import render
from collections import OrderedDict
from string import ascii_uppercase
from copy import deepcopy

CACHE_LIST_KEY = "liquidacion_list"
CACHE_DETAIL_PREFIX = "liquidacion_detail_"

# ─── Modelos y Serializers propios ─────────────────
from django.conf import settings
from datetime import date, datetime, timedelta
from django.utils.timezone import now
from .models import (
    DashboardCotizacion,
    DashboardOportunidad,
    vc_tab_areas,
    vc_tab_cargos,
    vc_tab_clientes,
    vc_tab_clientes_d,
    vc_tab_estado,
    vc_mov_cotizaciones,
    seg_usuario,
    cont_cias,
    vc_tab_clientes_d,
    CotiSuministros,
    CotiServicios,
    CotiMensajes,
    CotiSeguimiento,
    vc_tab_tproveedor,
    vc_tab_categorias,
    vc_tab_tgastos,
    vc_tab_tgastos_d,
    vc_tab_rittal,
    vc_tab_rockwell,
    vc_tab_ceyesa,
    vc_tab_hoffman,
    alm_articulos,
    ObjetivoAnualArea,
    ObjetivoAnual,
    Notificacion,
    vc_tab_notas,
    vc_mov_orden,
    )
from .serializers import (
    DashboardCotizacionTablaSerializer,
    DashboardOportunidadTablaSerializer,
    AreasSerializer,
    CargosSerializer,
    ClientesSerializer,
    RepresentantesSerializer,
    EstadoSerializer,
    CotizacionesSerializer,
    SegUsuarioSerializer,
    ContCiasSerializer,
    DashboardCotizacionModalSerializer,
    CotiSuministrosSerializer,
    CotiServiciosSerializer,
    CotiMensajesSerializer,
    CotiSeguimientoSerializer,
    DashboardCotizacionSerializer,
    ProveedoresSerializer,
    CategoriasSerializer,
    TGastosSerializer,
    TGastosDSerializer,
    RittalSerializer,
    RockwellSerializer,
    CeyesaSerializer,
    HoffmanSerializer,
    AlmArticulosSerializer,
    ObjetivoAnualAreaSerializer,
    ObjetivoAnualSerializer,
    NotificacionSerializer,
    NotasSerializer,
)

PLANTILLAS_DIR = os.path.join(os.path.dirname(__file__), "plantillas")

def siguiente_version(cotin):
    import re
    match = re.search(r'([A-Z])', cotin)
    if not match:
        raise ValueError("No se encontró versión en el cotin")

    letra_actual = match.group(1)
    nueva_letra = chr(ord(letra_actual) + 1)

    return cotin.replace(letra_actual, nueva_letra, 1)

# ===== Obtener y asegurar token CSRF =====
@ensure_csrf_cookie
def get_csrf_token(request):
    """
    Establece una cookie CSRF en el cliente. 
    Útil para peticiones POST protegidas desde el frontend.
    """
    return JsonResponse({'message': 'CSRF token set correctly.'}, status=200)

#========================================================================================

#=========#
# USUARIO #
#=========#
from cotizaciones_api.models import SegUsuario
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth.hashers import check_password
from cotizaciones_api.serializers import SegUsuarioSerializer
from rest_framework_simplejwt.authentication import JWTAuthentication

# Login DB_VC
@csrf_exempt
@api_view(['POST'])
def login_usuario(request):
    """
    Login seguro usando la tabla seg_usuarios (base empresarial).
    Usa usuario_usu como identificador único (user_id) en el JWT.
    Compatible con contraseñas planas o hasheadas.
    Devuelve JWT y datos del usuario.
    """
    usuario_input = request.data.get("usuario_usu")
    password_input = request.data.get("password_usu")

    # Validación básica
    if not usuario_input or not password_input:
        return Response(
            {"error": "Debe enviar usuario y contraseña."},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        # Buscar usuario en base principal
        usuario = SegUsuario.objects.using("default").get(usuario_usu=usuario_input.strip())
    except SegUsuario.DoesNotExist:
        return Response(
            {"error": "Usuario no encontrado."},
            status=status.HTTP_401_UNAUTHORIZED
        )

    # Validar contraseña (plana o hasheada)
    password_db = (usuario.password_usu or "").strip()
    if not (password_db == password_input or check_password(password_input, password_db)):
        return Response(
            {"error": "Contraseña incorrecta."},
            status=status.HTTP_401_UNAUTHORIZED
        )

    # ==========================
    #  GENERAR TOKEN PERSONALIZADO
    # ==========================
    refresh = RefreshToken()
    access = refresh.access_token

    # Usar usuario_usu como identificador único
    refresh["user_id"] = usuario.usuario_usu
    access["user_id"] = usuario.usuario_usu
    refresh["username"] = usuario.usuario_usu
    access["username"] = usuario.usuario_usu

    # (Opcional) incluir nombre corto o cargo para validaciones rápidas en frontend
    refresh["nombre"] = usuario.nomb_cort_usu
    access["nombre"] = usuario.nomb_cort_usu

    # Serializar datos del usuario
    user_data = SegUsuarioSerializer(usuario).data

    # Respuesta final
    return Response({
        "access": str(access),
        "refresh": str(refresh),
        "user": user_data
    }, status=status.HTTP_200_OK)

# Datos Usuario
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def usuario_actual(request):
    """
    Devuelve la información del usuario autenticado usando el JWT.
    Obtiene el usuario desde el claim user_id del token.
    """
    # Decodificar el token manualmente
    jwt_auth = JWTAuthentication()
    header = jwt_auth.get_header(request)
    if header is None:
        return Response({"error": "Token no proporcionado."}, status=status.HTTP_401_UNAUTHORIZED)

    raw_token = jwt_auth.get_raw_token(header)
    validated_token = jwt_auth.get_validated_token(raw_token)

    # Extraer user_id (que es usuario_usu)
    usuario_usu = validated_token.get("user_id")

    if not usuario_usu:
        return Response(
            {"error": "No se pudo obtener el usuario desde el token."},
            status=status.HTTP_401_UNAUTHORIZED
        )

    try:
        usuario = seg_usuario.objects.using("default").get(usuario_usu=usuario_usu)
    except seg_usuario.DoesNotExist:
        return Response(
            {"error": "Usuario no encontrado en la base de datos."},
            status=status.HTTP_404_NOT_FOUND
        )

    user_data = SegUsuarioSerializer(usuario).data
    return Response(user_data, status=status.HTTP_200_OK)

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def usuarios_activos(request):
    """
    Devuelve usuarios activos (activo = 1)
    Permite búsqueda por:
    - DNI
    - Email
    - Nombre corto
    - Usuario
    """
    q = request.GET.get("q", "").strip()

    usuarios = seg_usuario.objects.using("default").filter(activo=1)

    if q:
        usuarios = usuarios.filter(
            Q(dni__icontains=q) |
            Q(email_usu__icontains=q) |
            Q(nomb_cort_usu__icontains=q) |
            Q(usuario_usu__icontains=q)
        )

    serializer = SegUsuarioSerializer(usuarios, many=True)
    return Response(serializer.data, status=status.HTTP_200_OK)

#========================================================================================

#=========================#
# APROBACION COTIZACIONES #
#=========================#
# ─── Dashboard Cotizaciones (versión moderna) ─────────────────────
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def cotizaciones_dashboard_view(request):
    """
    Dashboard + Tabla para cotizaciones con filtros flexibles.
    Estilo moderno equivalente al dashboard de CAJA CHICA.
    """
    try:
        from datetime import date
        from unidecode import unidecode
        from django.db.models import Func, F, Value, TextField
        from django.db.models.functions import Lower
        import re

        # ============================================================
        # 1) Parámetros principales
        # ============================================================
        anno = request.GET.get("anno", date.today().year)
        mes = request.GET.get("mes", "%")

        cliente = request.GET.get("cliente", "%")
        estado = request.GET.get("estado", "%")
        area = request.GET.get("area", "%")
        envio = request.GET.get("envio", "%")

        CAMPOS_BUSQUEDA = {
            "num_reg": "num_reg",
            "cotin": "numero",
            "cotif": "fecha",
            "cliente_nombre": "cliente_nombre",
            "refef": "referencia",
            "nombr": "nombr",
            "nombc": "nombc",
            "nombt": "nombt",
            "tot_c": "tot_c",
            "tot_d": "tot_d",
            "prob": "prob",
            "regus": "regus",
        }

        # Búsqueda flexible
        campo = request.GET.get("campo")
        valor = request.GET.get("valor")

        fecha_inicio = request.GET.get("fechaInicio")
        fecha_fin = request.GET.get("fechaFin")

        # ============================================================
        # 2) Query base
        # ============================================================
        qs = DashboardCotizacion.objects.all()

        if anno != "%":
            qs = qs.filter(anno_a=anno)

        if mes != "%":
            qs = qs.filter(fecha__month=mes)

        if cliente != "%":
            qs = qs.filter(cliente_codigo=cliente)

        if estado != "%":
            estados = [e for e in estado.split(",") if e]

            if len(estados) == 1:
                qs = qs.filter(estado_codigo=estados[0])
            else:
                qs = qs.filter(estado_codigo__in=estados)

        if area != "%":
            qs = qs.filter(area_codigo=area)

        if envio != "%":
            qs = qs.filter(envio=envio)

        # Rango de fechas
        if fecha_inicio:
            qs = qs.filter(fecha__gte=fecha_inicio)
        if fecha_fin:
            qs = qs.filter(fecha__lte=fecha_fin)

        # ============================================================
        # 3) Normalización de búsqueda flexible
        # ============================================================
        def normalizar(texto):
            if not texto:
                return None
            t = unidecode(texto.lower().strip())
            return re.sub(r"\s+", " ", t)

        class Replace(Func):
            function = "REPLACE"
            arity = 3

        if campo and valor not in (None, "", " "):
            campo_real = CAMPOS_BUSQUEDA.get(campo)

            if campo_real:
                valor_norm = normalizar(valor)

                qs = qs.annotate(
                    campo_clean=Replace(
                        Replace(
                            Lower(F(campo_real)),
                            Value("  ", output_field=TextField()),
                            Value(" ", output_field=TextField()),
                            output_field=TextField()
                        ),
                        Value("  ", output_field=TextField()),
                        Value(" ", output_field=TextField()),
                        output_field=TextField()
                    )
                ).filter(campo_clean__icontains=valor_norm)

        # ============================================================
        # 4) Dashboard Stats mejorado
        # ============================================================
        total = qs.count()

        estados_db = vc_tab_estado.objects.filter(activo=True).values_list("nombre", flat=True)
        estado_map = {e: 0 for e in estados_db}

        meses = [0] * 12
        monto_total_soles = 0
        monto_total_dolares = 0
        este_mes = 0

        hoy = date.today()

        # Diccionario para stats por cliente
        clientes_stats = {}

        for c in qs:
            # ===== Estados =====
            estado_nombre = c.estado_nombre or "Pendiente"
            estado_map[estado_nombre] = estado_map.get(estado_nombre, 0) + 1

            # ===== Monto por moneda =====
            if hasattr(c, "tmone"):
                if c.tmone == "S":
                    monto_total_soles += float(c.tot_c or 0)
                elif c.tmone == "D":
                    monto_total_dolares += float(c.tot_c or 0)
            else:
                # si no hay tmone definido, asumimos S/ por defecto
                monto_total_soles += float(c.tot_c or 0)

            # ===== Conteo por mes =====
            if c.fecha:
                idx = c.fecha.month - 1
                meses[idx] += 1
                if c.fecha.month == hoy.month:
                    este_mes += 1

            # ===== Stats por cliente =====
            codigo = c.cliente_codigo  # código del cliente
            nombre = c.cliente_nombre or "-"
            
            if codigo not in clientes_stats:
                clientes_stats[codigo] = {
                    "cliente_codigo": codigo,
                    "nombre": nombre,
                    "cantidad": 0,
                    "totalSoles": 0,
                    "totalDolares": 0,
                }

            clientes_stats[codigo]["cantidad"] += 1
            if hasattr(c, "tmone"):
                if c.tmone == "S":
                    clientes_stats[codigo]["totalSoles"] += float(c.tot_c or 0)
                elif c.tmone == "D":
                    clientes_stats[codigo]["totalDolares"] += float(c.tot_c or 0)
            else:
                clientes_stats[codigo]["totalSoles"] += float(c.tot_c or 0)

        # Calcular porcentaje de uso por cliente
        for cliente in clientes_stats.values():
            cliente["porcentaje"] = round((cliente["cantidad"] / total) * 100, 2) if total else 0

        # Convertir a lista para enviar al frontend
        clientes_list = list(clientes_stats.values())

        # ===== Dashboard final =====
        dashboard_data = {
            "total": total,
            "esteMes": este_mes,
            "montoTotalSoles": round(monto_total_soles, 2),
            "montoTotalDolares": round(monto_total_dolares, 2),
            "promedioSoles": round(monto_total_soles / total, 2) if total else 0,
            "promedioDolares": round(monto_total_dolares / total, 2) if total else 0,
            "estados": estado_map,
            "porMes": meses,
            "clientes": clientes_list,  # ✅ aquí agregamos los stats de clientes
        }

        # ============================================================
        # 5) Tabla de registros
        # ============================================================
        tabla_data = DashboardCotizacionTablaSerializer(
            qs.order_by(
                F("envio").asc(nulls_last=True),
                F("fecha").desc(),
                F("num_reg").desc(),
            ),
            many=True
        ).data

        # ============================================================
        # 6) Respuesta final
        # ============================================================
        return Response({
            "dashboard": dashboard_data,
            "tabla": tabla_data,
            "anno": anno
        })

    except Exception:
        import traceback
        print(traceback.format_exc())
        return Response({"error": "Error interno en el servidor."}, status=500)

# Detalle de cotización por num_reg
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def cotizacion_modal_view(request, num_reg):
    """
    Retorna los detalles completos de una cotización usando num_reg 
    (clave principal real del registro).
    """
    try:
        # Buscar la cotización por num_reg
        cot = DashboardCotizacion.objects.filter(num_reg=num_reg).first()
        if not cot:
            return Response(
                {"error": f"No se encontró la cotización con num_reg {num_reg}"},
                status=404
            )

        # Serializar
        serializer = DashboardCotizacionModalSerializer(cot)
        return Response(serializer.data)

    except Exception as e:
        import traceback
        print("Error en cotizacion_modal_view:", traceback.format_exc())
        return Response({"error": str(e)}, status=500)

@api_view(["GET", "POST", "PUT"])
@permission_classes([IsAuthenticated])
def listar_suministros(request, num_reg):
    try:
        # ======================
        # 📄 LISTAR
        # ======================
        if request.method == "GET":
            suministros = CotiSuministros.objects.filter(
                num_reg=num_reg
            ).order_by("num")

            serializer = CotiSuministrosSerializer(suministros, many=True)
            return Response(serializer.data)

        # ======================
        # ➕ CREAR
        # ======================
        if request.method == "POST":
            data = request.data.copy()
            data["num_reg"] = num_reg

            serializer = CotiSuministrosSerializer(data=data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=201)

            return Response(serializer.errors, status=400)

        # ======================
        # ✏️ ACTUALIZAR
        # ======================
        if request.method == "PUT":
            item_id = request.data.get("id")

            try:
                item_id = int(item_id)
            except (TypeError, ValueError):
                return Response({"error": "ID inválido"}, status=400)

            suministro = CotiSuministros.objects.get(
                id=item_id,
                num_reg=num_reg
            )

            data = request.data.copy()
            data.pop("id", None)
            data.pop("num_reg", None)

            serializer = CotiSuministrosSerializer(
                suministro,
                data=data,
                partial=True
            )

            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)

            return Response(serializer.errors, status=400)

    except CotiSuministros.DoesNotExist:
        return Response({"error": "Suministro no encontrado"}, status=404)

    except Exception as e:
        return Response({"error": str(e)}, status=500)

def clean_text(text):
    """Limpia espacios, tabulaciones, saltos de línea y decodifica HTML."""
    if not text:
        return ""
    # Decodifica entidades HTML
    text = unescape(text)
    # Reemplaza cualquier secuencia de espacios o saltos de línea por un solo espacio
    text = re.sub(r"\s+", " ", text)
    return text.strip()

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def listar_servicios(request, num_reg):
    try:
        # 1) OBTENER TODOS LOS SERVICIOS (nig = 0)
        servicios_qs = CotiServicios.objects.filter(
            num_reg=num_reg,
            nig=0
        ).order_by("num")

        resultado = []

        tipo_map = {
            "4": "MANO DE OBRA",
            "5": "GASTOS DE SERVICIOS",
            "6": "OTROS",
        }

        for servicio in servicios_qs:
            servicio_data = CotiServiciosSerializer(servicio).data
            titulo_general = clean_text(servicio_data.get("nog"))

            pref_servicio = servicio.cog[:2]

            # 2) FILAS DEL SERVICIO (subgrupos + items)
            rows = CotiServicios.objects.filter(
                num_reg=num_reg,
                cog__startswith=pref_servicio,
                nig__gt=0
            ).order_by("num")

            subgrupos = []

            # 3) CREAR SUBGRUPOS (nig = 1)
            for row in rows.filter(nig=1):
                tipo_digito = row.cog[3]  # 4,5,6
                subgrupo = {
                    "titulo": clean_text(row.nog),       # 🔹 título real de DB
                    "tipoCodigo": f"0{tipo_digito}",     # "04", "05", "06"
                    "tipoNombre": tipo_map.get(tipo_digito, "DESCONOCIDO"),
                    "items": []
                }
                subgrupos.append(subgrupo)

            # 4) ASIGNAR ITEMS (nig = 2)
            for row in rows.filter(nig=2):
                # Buscar el subgrupo correspondiente según los primeros 4 dígitos del cog
                prefijo = row.cog[:4]
                for sg in subgrupos:
                    if sg["tipoCodigo"] == f"0{row.cog[3]}":
                        item = CotiServiciosSerializer(row).data
                        for f in ["des", "cod", "nog", "pro"]:
                            item[f] = clean_text(item.get(f))
                        sg["items"].append(item)
                        break

            resultado.append({
                "tituloGeneral": titulo_general,
                "cantidad": str(servicio.can or "1"),   # 🔹 can (nig = 0)
                "detalle": servicio.tog or "",           # 🔹 tog (HTML)
                "subgrupos": subgrupos
            })

        return Response(resultado)

    except Exception as e:
        return Response({"error": str(e)}, status=500)

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def listar_mensajes(request, num_reg):
    """
    Lista todos los mensajes asociados a un num_reg.
    """
    try:
        mensajes = CotiMensajes.objects.filter(
            num_reg=num_reg,
            act="1"  # Solo activos
        ).order_by("dat")  # Orden cronológico, más antiguos primero

        serializer = CotiMensajesSerializer(mensajes, many=True)
        return Response(serializer.data)

    except Exception as e:
        return Response({"error": str(e)}, status=500)

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def listar_seguimientos(request, num_reg):
    """
    Lista todos los seguimientos asociados a un num_reg.
    """
    try:
        seguimientos = CotiSeguimiento.objects.filter(
            num_reg=num_reg,
            act="1"  # Solo activos
        ).order_by("dat")  # Orden cronológico, más antiguos primero

        serializer = CotiSeguimientoSerializer(seguimientos, many=True)
        return Response(serializer.data)

    except Exception as e:
        return Response({"error": str(e)}, status=500)

@api_view(["GET"])
def totales_descuento_view(request, num_reg):
    cot = DashboardCotizacion.objects.get(num_reg=num_reg)

    total = cot.tot_c or 0
    descuento_monto = cot.des_m or 0  # <-- descuento guardado

    total_suministros = (
        CotiSuministros.objects
        .filter(num_reg=num_reg, nig=0)
        .aggregate(
            total=Sum(F("tot") * F("can"))
        )["total"] or 0
    )

    total_servicios = (
        CotiServicios.objects
        .filter(num_reg=num_reg, nig=0)
        .aggregate(
            total=Sum(F("tot") * F("can"))
        )["total"] or 0
    )

    return Response({
        "total": total,
        "suministros": total_suministros,
        "servicios": total_servicios,
        "tot_c": total,             # total base
        "des_m": descuento_monto,   # descuento guardado
    })

@api_view(["PATCH"])
@permission_classes([IsAuthenticated])
def recalcular_totales_cotizacion(request, num_reg):
    """
    Recalcula el total de la cotización basado en suministros y servicios.
    Retorna el total actualizado.
    """
    try:
        cotizacion = DashboardCotizacion.objects.get(num_reg=num_reg)

        # Totales suministros
        total_suministros = (
            CotiSuministros.objects
            .filter(num_reg=num_reg, nig=0)
            .aggregate(total=Sum(F("tot") * F("can")))["total"] or 0
        )

        # Totales servicios
        total_servicios = (
            CotiServicios.objects
            .filter(num_reg=num_reg, nig=0)
            .aggregate(total=Sum(F("tot") * F("can")))["total"] or 0
        )

        total_general = total_suministros + total_servicios

        # Guardar total
        cotizacion.tot_c = total_general
        cotizacion.save(update_fields=["tot_c"])

        return Response({
            "tot_c": total_general,
            "suministros": total_suministros,
            "servicios": total_servicios,
        })

    except DashboardCotizacion.DoesNotExist:
        return Response({"error": "Cotización no encontrada"}, status=404)

#========================================================================================

##===============##
## OPORTUNIDADES ##
##===============##
# ─── Dashboard Cotizaciones (versión moderna) ─────────────────────
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def oportunidades_dashboard_view(request):
    """
    Dashboard + Tabla para Oportunidades con filtros flexibles.
    Sincronizado con la estructura de vc_mov_oportunidades.
    """
    try:
        from datetime import date
        from unidecode import unidecode
        from django.db.models import Func, F, Value, TextField
        from django.db.models.functions import Lower
        import re

        # ============================================================
        # 1) Parámetros principales
        # ============================================================
        anno = request.GET.get("anno", date.today().year)
        mes = request.GET.get("mes", "%")

        cliente = request.GET.get("cliente", "%")
        estado = request.GET.get("estado", "%")
        area = request.GET.get("area", "%")
        responsable = request.GET.get("responsable", "%")

        # Mapeo de búsqueda flexible para Oportunidades
        CAMPOS_BUSQUEDA = {
            "num_reg": "num_reg",
            "codig": "codig",
            "f_recp": "f_recp",
            "cliente_nombre": "nombr",
            "contac": "contac",
            "descr": "descr",
            "respo": "respo",
            "monto": "monto",
            "comen": "comen",
            "regus": "regus",
        }

        campo = request.GET.get("campo")
        valor = request.GET.get("valor")

        fecha_inicio = request.GET.get("fechaInicio") # Basado en f_recp
        fecha_fin = request.GET.get("fechaFin")

        # ============================================================
        # 2) Query base
        # ============================================================
        qs = DashboardOportunidad.objects.filter(
            anno_a=anno
        )

        if mes != "%":
            qs = qs.filter(f_recp__month=mes)

        if cliente != "%":
            qs = qs.filter(empre=cliente)

        if estado != "%":
            estados = [e for e in estado.split(",") if e]
            if len(estados) == 1:
                qs = qs.filter(estad=estados[0])
            else:
                qs = qs.filter(estad__in=estados)

        if area != "%":
            qs = qs.filter(area=area)

        if responsable != "%":
            qs = qs.filter(respo=responsable)

        # Rango de fechas (sobre fecha de recepción)
        if fecha_inicio:
            qs = qs.filter(f_recp__gte=fecha_inicio)
        if fecha_fin:
            qs = qs.filter(f_recp__lte=fecha_fin)

        # ============================================================
        # 3) Normalización de búsqueda flexible
        # ============================================================
        def normalizar(texto):
            if not texto: return None
            t = unidecode(texto.lower().strip())
            return re.sub(r"\s+", " ", t)

        class Replace(Func):
            function = "REPLACE"
            arity = 3

        if campo and valor not in (None, "", " "):
            campo_real = CAMPOS_BUSQUEDA.get(campo)
            if campo_real:
                valor_norm = normalizar(valor)
                qs = qs.annotate(
                    campo_clean=Replace(
                        Replace(
                            Lower(F(campo_real)),
                            Value("  ", output_field=TextField()),
                            Value(" ", output_field=TextField()),
                            output_field=TextField()
                        ),
                        Value("  ", output_field=TextField()),
                        Value(" ", output_field=TextField()),
                        output_field=TextField()
                    )
                ).filter(campo_clean__icontains=valor_norm)

        # ============================================================
        # 4) Dashboard Stats Oportunidades
        # ============================================================
        total = qs.count()
        
        # Mapeo de estados manual para el conteo de stats
        estado_map = {"PENDIENTE": 0, "COTIZADO": 0, "PERDIDO": 0, "ADJUDICADO": 0}

        meses = [0] * 12
        monto_total_soles = 0
        monto_total_dolares = 0
        este_mes = 0
        hoy = date.today()
        clientes_stats = {}

        for o in qs:
            # ===== Estados =====
            est_nom = o.estado_nombre
            estado_map[est_nom] = estado_map.get(est_nom, 0) + 1

            # ===== Montos =====
            val_monto = float(o.monto or 0)
            if o.tmone == "S":
                monto_total_soles += val_monto
            else: # Dólares por defecto
                monto_total_dolares += val_monto

            # ===== Conteo temporal =====
            if o.f_recp:
                idx = o.f_recp.month - 1
                meses[idx] += 1
                if o.f_recp.month == hoy.month:
                    este_mes += 1

            # ===== Stats por cliente =====
            cod_cli = o.empre or "VAR"
            nom_cli = o.nombr or "CLIENTE VARIO"
            
            if cod_cli not in clientes_stats:
                clientes_stats[cod_cli] = {
                    "cliente_codigo": cod_cli,
                    "nombre": nom_cli,
                    "cantidad": 0,
                    "totalSoles": 0,
                    "totalDolares": 0,
                }
            clientes_stats[cod_cli]["cantidad"] += 1
            if o.tmone == "S":
                clientes_stats[cod_cli]["totalSoles"] += val_monto
            else:
                clientes_stats[cod_cli]["totalDolares"] += val_monto

        # Porcentajes por cliente
        for c_data in clientes_stats.values():
            c_data["porcentaje"] = round((c_data["cantidad"] / total) * 100, 2) if total else 0

        dashboard_data = {
            "total": total,
            "esteMes": este_mes,
            "montoTotalSoles": round(monto_total_soles, 2),
            "montoTotalDolares": round(monto_total_dolares, 2),
            "promedioSoles": round(monto_total_soles / total, 2) if total else 0,
            "promedioDolares": round(monto_total_dolares / total, 2) if total else 0,
            "estados": estado_map,
            "porMes": meses,
            "clientes": list(clientes_stats.values()),
        }

        # ============================================================
        # 5) Serialización de Tabla
        # ============================================================
        # Nota: Debes crear este Serializer similar al de cotizaciones
        tabla_data = DashboardOportunidadTablaSerializer(
            qs.order_by("-f_recp", "-num_reg"),
            many=True
        ).data

        return Response({
            "dashboard": dashboard_data,
            "tabla": tabla_data,
            "anno": anno
        })

    except Exception:
        import traceback
        print(traceback.format_exc())
        return Response({"error": "Error interno en el servidor de Oportunidades."}, status=500)

#========================================================================================

##=========##
## GUARDAR ##
##=========##
@csrf_exempt
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def guardar_cotizacion(request):
    with transaction.atomic():
        # Capturamos el num_reg que viene del frontend
        num_reg_frontend = request.data.get("num_reg")
        cotizacion = None

        # =========================
        # 1️⃣ BUSCAR O CREAR
        # =========================
        if num_reg_frontend:
            # Intentamos buscar si ya existe para ACTUALIZAR
            cotizacion = DashboardCotizacion.objects.filter(num_reg=num_reg_frontend).select_for_update().first()

        if not cotizacion:
            # SI NO EXISTE: Es una creación nueva.
            nuevo_num = num_reg_frontend if num_reg_frontend else obtener_siguiente_num_reg()
            hoy = timezone.now()
            
            # 💡 IMPORTANTE: Pasamos anno_a y campos críticos directamente en el .create()
            # para evitar que MySQL rechace el registro por restricciones NOT NULL.
            cotizacion = DashboardCotizacion.objects.create(
                num_reg=nuevo_num,
                fecha=request.data.get("fecha", hoy.date()),
                envio=0,
                sald=Decimal("0.00"),
                tot_c=Decimal("0.00"),
                igv="N",
                anno_a=str(hoy.year),  # <-- Esto arregla tu error de DB
                anno=str(hoy.year),
                mes=str(hoy.month).zfill(2)
            )
            es_creacion = True
        else:
            es_creacion = False

        # =========================
        # 2️⃣ ACTUALIZAR DATOS (EXCEPTO num_reg)
        # =========================
        data = request.data.copy()
        
        data.pop("num_reg", None)

        if "acu_e" not in data:
            data.pop("acu_e", None)

        serializer = DashboardCotizacionSerializer(
            cotizacion,
            data=data,
            partial=True
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()

        cotizacion.igv = cotizacion.igv or "N"

        # =========================
        # 3️⃣ SNAPSHOT CLIENTE
        # =========================
        if cotizacion.cliente_codigo:
            cliente = vc_tab_clientes_d.objects.filter(
                codigo=cotizacion.cliente_codigo
            ).first()

            if cliente:
                cotizacion.codir = cliente.codigo
                cotizacion.nombr = cliente.representante
                cotizacion.cargr = cliente.cargo
                cotizacion.teler = cliente.telefono
                cotizacion.movir = cliente.movil
                cotizacion.mailr = cliente.email

        if cotizacion.fecha:
            cotizacion.anno = str(cotizacion.fecha.year).zfill(4)
            cotizacion.mes = str(cotizacion.fecha.month).zfill(2)

        usuario = seg_usuario.objects.filter(
            nomb_cort_usu=cotizacion.nombc
        ).first()
        cotizacion.codic = usuario.dni if usuario else None

        cotizacion.save()

        # =========================
        # 4️⃣ SUMINISTROS
        # =========================
        suministros = request.data.get("suministros", {})
        
        # 💡 Usamos el campo tven del modelo DashboardCotizacion
        # T = Venta Total, P = Venta Parcial
        tipo_venta_general = cotizacion.tven 

        CotiSuministros.objects.filter(num_reg=cotizacion.num_reg).delete()

        TIPO_MAP = {"01": "01", "02": "02"}
        num_contador = 1
        grupo_index = 1

        for _, grupo in suministros.items():
            cog = grupo.get("cog") or grupo.get("id")
            if not cog:
                tipo = grupo.get("tipo", "01")
                tipo_code = TIPO_MAP.get(tipo, "01")
                cog = f"{grupo_index:02d}{tipo_code}"
                grupo_index += 1

            cantidad_grupo = Decimal(str(grupo.get("cantidad", 0)))
            total_grupo = Decimal(str(grupo.get("total", 0)))
            
            # Valor que viene del modal de grupo (React)
            costo_envio_valor = Decimal(str(grupo.get("costoEnvio", 0)))

            # =====================
            # CABECERA (nig = 0)
            # =====================
            CotiSuministros.objects.create(
                num_reg=cotizacion.num_reg,
                cog=cog,
                nog=grupo.get("titulo"),
                nig=0,
                num=num_contador,
                cod="0",
                can=cantidad_grupo,
                tot=total_grupo,

                # 💡 Lógica basada en tven:
                # Si tven es 'T' -> guarda en env_tot
                # Si tven es 'P' -> guarda en env_par (Parcial)
                env_tot=costo_envio_valor if tipo_venta_general == "T" else Decimal("0.00"),
                env_par=costo_envio_valor if tipo_venta_general == "P" else Decimal("0.00"),
                cost_c_env=Decimal("0.00"),

                mov="01",
                tog="0",
                cost_env=Decimal("0.00"),
                por_env=Decimal("0.00"),
            )

            num_contador += 1

            # =====================
            # ITEMS (nig = 1)
            # =====================
            for item in grupo.get("items", []):
                CotiSuministros.objects.create(
                    num_reg=cotizacion.num_reg,
                    cog=cog,
                    nog="",
                    nig=1,
                    num=num_contador,
                    cod=item.get("cod"),
                    des=item.get("des"),
                    pro=item.get("pro"),
                    can=Decimal(str(item.get("can", 0))),
                    puc=Decimal(str(item.get("puc", 0))),
                    toc=Decimal(str(item.get("toc", 0))),
                    cau=Decimal(str(item.get("cau", 0))),
                    tou=Decimal(str(item.get("tou", 0))),
                    val=Decimal(str(item.get("val", 0))),
                    tot=Decimal(str(item.get("tot", 0))),
                    
                    # Costos unitarios por ítem
                    cost_env=Decimal(str(item.get("cost_env", 0))),
                    por_env=Decimal(str(item.get("por_env", 0))),
                    
                    env_tot=Decimal("0.00"),
                    env_par=Decimal("0.00"),
                    
                    mov="01",
                    tpr=item.get("tpr"),
                    tde=item.get("tde"),
                    tog="0",
                    ent=item.get("ent"),
                    enu=item.get("enu"),
                    obs=item.get("obs"),
                )
                num_contador += 1
                
        # =========================
        # 5️⃣ SERVICIOS
        # =========================
        servicios = request.data.get("servicios", {})

        if isinstance(servicios, list):
            servicios = {str(i): s for i, s in enumerate(servicios)}

        # Limpieza total (igual que suministros)
        CotiServicios.objects.filter(
            num_reg=cotizacion.num_reg
        ).delete()

        SUB_MAP = {
            "MANO DE OBRA": "04",
            "GASTOS DE SERVICIO": "05",
            "OTROS": "06",
        }

        for _, servicio in servicios.items():

            # =====================
            # CABECERA SERVICIO (nig = 0)
            # =====================
            cog_servicio = f"{grupo_index:02d}000"
            total_servicio = Decimal("0.00")
            cantidad_servicio = Decimal(str(servicio.get("cantidad", 1)))

            CotiServicios.objects.create(
                num_reg=cotizacion.num_reg,
                cog=cog_servicio,
                nog=servicio.get("tituloGeneral") or servicio.get("nombre", ""),
                nig=0,
                num=num_contador,
                cod="0",
                can=cantidad_servicio,
                tot=Decimal("0.00"),
                tog=servicio.get("detalle", ""),  # 🔴 AQUÍ
            )

            num_servicio = num_contador
            num_contador += 1

            # =====================
            # SUBGRUPOS (nig = 1)
            # =====================
            for sub in servicio.get("subgrupos", []):
                tipo = sub.get("titulo", "OTROS")
                tipo_code = sub.get("tipoCodigo", "06")  # 🔹 usamos el tipo que envía frontend
                total_sub = Decimal("0.00")

                cog_sub = f"{grupo_index:02d}{tipo_code}1"

                CotiServicios.objects.create(
                    num_reg=cotizacion.num_reg,
                    cog=cog_sub,
                    nog=tipo,
                    nig=1,
                    num=num_contador,
                    cod="0",
                    can=Decimal("0.00"),
                    tot=Decimal("0.00"),
                    mov=sub.get("mov", ""),
                )

                num_sub = num_contador
                num_contador += 1

                # =====================
                # ITEMS (nig = 2)
                # =====================
                for item in sub.get("items", []):
                    total_item = Decimal(str(item.get("tot", 0)))
                    cog_item = f"{grupo_index:02d}{tipo_code}2"

                    CotiServicios.objects.create(
                        num_reg=cotizacion.num_reg,
                        cog=cog_item,
                        nog="",
                        nig=2,
                        num=num_contador,
                        cod=(
                            f"{item.get('personalCodigo','')} - {item.get('personal','')}"
                            if tipo == "MANO DE OBRA"
                            else item.get("cod", "")
                        ),
                        des=item.get("des", ""),
                        pro=item.get("pro", 8),
                        can=Decimal(str(item.get("can", 1))),
                        puc=Decimal(str(item.get("puc", 0))),
                        toc=Decimal(str(item.get("toc", 0))),
                        cau=Decimal(str(item.get("cau", 0))),
                        tou=Decimal(str(item.get("tou", 0))),
                        val=Decimal(str(item.get("val", 0))),
                        tot=total_item,
                        mov=item.get("mov", ""),
                        tpr=item.get("tpr", ""),
                        tde=item.get("tde", 1),
                    )

                    total_sub += total_item
                    total_servicio += total_item
                    num_contador += 1

                # actualizar subtotal
                CotiServicios.objects.filter(
                    num_reg=cotizacion.num_reg,
                    num=num_sub,
                    nig=1
                ).update(tot=total_sub)

            # actualizar total servicio
            CotiServicios.objects.filter(
                num_reg=cotizacion.num_reg,
                num=num_servicio,
                nig=0
            ).update(tot=total_servicio)

            grupo_index += 1

        # =========================
        # 6️⃣ MENSAJES
        # =========================
        mensaje_data = request.data.get("mensaje")

        if mensaje_data:
            CotiMensajes.objects.create(
                num_reg=cotizacion.num_reg,
                cod=request.user.usuario_usu if hasattr(request.user, "usuario_usu") else request.user.username,
                msj=mensaje_data.get("msj"),
                act=mensaje_data.get("act", "1"),
            )

        # =========================
        # 7️⃣ SEGUIMIENTOS
        # =========================
        seguimiento_data = request.data.get("seguimiento")  # 🔑 similar a "mensaje"

        if seguimiento_data:
            # Crear nuevo registro en CotiSeguimiento
            CotiSeguimiento.objects.create(
                num_reg=cotizacion.num_reg,
                num=seguimiento_data.get("num", 0),  # opcional, si no lo envías
                dat=timezone.now(),  # se usa la fecha actual como PK virtual
                fec=timezone.now().strftime("%Y-%m-%d %H:%M:%S"),  # tu campo 'fec' string
                hor=timezone.now().strftime("%H:%M:%S"),  # opcional
                des=seguimiento_data.get("des"),
                cod=request.user.usuario_usu if hasattr(request.user, "usuario_usu") else request.user.username,
                act=seguimiento_data.get("act", "1"),
            )

        # =========================
        # 8️⃣ TOTAL GENERAL + DESCUENTO
        # =========================
        descuento = request.data.get("descuento", {})
        aplicar = descuento.get("aplicar", False)
        aplica_a = descuento.get("aplicaA")
        importe_desc = Decimal(str(descuento.get("importe", 0)))
        porcentaje_desc = Decimal(str(descuento.get("porcentaje", 0)))

        # Totales base
        total_suministros = sum(
            Decimal(str(grupo.get("total", 0))) *
            Decimal(str(grupo.get("cantidad", 1)))
            for grupo in suministros.values()
        )

        total_servicios = sum(
            Decimal(str(grupo.get("total", 0)))
            for grupo in servicios.values()
        )

        total_general = total_suministros + total_servicios

        # Aplicar descuento
        if aplicar and importe_desc > 0:
            if aplica_a == "TOTAL":
                total_general -= importe_desc
            elif aplica_a == "SUMINISTROS":
                total_general = (total_suministros - importe_desc) + total_servicios
            elif aplica_a == "SERVICIOS":
                total_general = total_suministros + (total_servicios - importe_desc)

        if total_general < 0:
            total_general = Decimal("0.00")

        # 👉 ASIGNAR TODO
        cotizacion.tot_c = total_general
        cotizacion.des_a = 1 if aplicar else 0

        if aplica_a == "TOTAL":
            cotizacion.des_t = "T"
        elif aplica_a == "SUMINISTROS":
            cotizacion.des_t = "S"
        elif aplica_a == "SERVICIOS":
            cotizacion.des_t = "M"
        else:
            cotizacion.des_t = None

        cotizacion.des_p = porcentaje_desc
        cotizacion.des_m = importe_desc

        # 👉 AHORA SÍ guardar
        cotizacion.save(update_fields=[
            "tot_c", "des_a", "des_t", "des_p", "des_m",
        ])

        return Response(
            {
                "message": "Cotización guardada correctamente",
                "num_reg": cotizacion.num_reg
            },
            # Ajustamos la condición aquí:
            status=status.HTTP_200_OK if not es_creacion else status.HTTP_201_CREATED
        )

# ADICIONAL
def obtener_siguiente_num_reg():
    anio = timezone.now().year

    inicio = int(f"{anio}000000")
    fin = int(f"{anio}999999")

    with transaction.atomic():
        ultimo = (
            DashboardCotizacion.objects
            .select_for_update()
            .filter(num_reg__range=(inicio, fin))
            .aggregate(max_reg=Max("num_reg"))
        )["max_reg"]

        if ultimo is not None:
            return ultimo + 1

        # 🚀 Primer registro del año
        return int(f"{anio}000001")

#========================================================================================

##===========##
## BUSQUEDAS ##
##===========##
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def buscar_encargados_por_empresa(request, empresa):
    q = request.GET.get("q", "").strip()

    # Filtramos en vc_tab_clientes_d
    encargados = vc_tab_clientes_d.objects.filter(
        empresa=empresa,    # El campo 'empresa' de la DB coincide con el ID del cliente
        activo="1"          # IMPORTANTE: En tu modelo es CharField, usamos "1" no True
    ).filter(
        Q(representante__icontains=q) |
        Q(codigo__icontains=q)
    ).values(
        "codigo",
        "representante",
        "cargo",
        "telefono",
        "movil",
        "email",
        "empresa"
    )

    return Response(list(encargados))

#========================================================================================
##==========##
## ADJUNTOS ##
##==========##
@csrf_exempt
def subir_archivo(request):
    if request.method == "POST":
        print("FILES recibidos:", request.FILES)
        archivo = request.FILES.get("archivo")
        nombre_guardar = request.POST.get("nombre")  # <--- esto es lo nuevo

        if archivo:
            ruta = r"C:\Users\VC-23031\PROYECTOS\Adj"
            try:
                # usa el nombre que enviaste desde React
                with open(os.path.join(ruta, nombre_guardar), "wb+") as destino:
                    for chunk in archivo.chunks():
                        destino.write(chunk)
                return JsonResponse({"ok": True, "archivo": nombre_guardar})
            except Exception as e:
                print("Error al escribir archivo:", e)
                return JsonResponse({"ok": False, "error": str(e)})
        else:
            return JsonResponse({"ok": False, "error": "No se recibió archivo"})
    return JsonResponse({"ok": False, "error": "Método no permitido"})

@csrf_exempt
def listar_adjuntos(request, num_reg):
    carpeta = r"C:\Users\VC-23031\PROYECTOS\Adj"
    try:
        archivos = []
        # Itera los archivos de la carpeta
        for nombre in os.listdir(carpeta):
            if nombre.startswith(str(num_reg)):  # solo archivos que empiezan con num_reg
                archivos.append({
                    "nombre": nombre,               # nombre completo guardado en disco
                    "displayName": nombre[len(str(num_reg)) + 1:],  # quitar prefijo para mostrar
                })
        return JsonResponse({"ok": True, "archivos": archivos})
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)})

@csrf_exempt
def eliminar_archivo(request):
    if request.method == "POST":
        nombre = request.POST.get("nombre")

        if not nombre:
            return JsonResponse({"ok": False, "error": "Nombre no recibido"})

        ruta = r"C:\Users\VC-23031\PROYECTOS\Adj"
        path = os.path.join(ruta, nombre)

        if not os.path.exists(path):
            return JsonResponse({"ok": False, "error": "Archivo no existe"})

        try:
            os.remove(path)
            return JsonResponse({"ok": True})
        except Exception as e:
            return JsonResponse({"ok": False, "error": str(e)})

    return JsonResponse({"ok": False, "error": "Método no permitido"})

#========================================================================================

##=========##
## GESTION ##
##=========##
@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def condiciones_generales(request, num_reg):
    """
    GET:  retorna {"numero", "condiciones"}  (acu_e)
    POST: recibe { "condiciones": "texto..." } y actualiza acu_e
    """
    try:
        cot = DashboardCotizacion.objects.filter(num_reg=num_reg).first()
        if not cot:
            return Response({"error": f"No se encontró la cotización #{num_reg}"}, status=404)

        if request.method == "GET":
            return Response({
                "numero": cot.num_reg,
                "condiciones": cot.acu_e or ""
            }, status=200)

        # POST -> actualizar
        contenido = request.data.get("condiciones", "")
        # opcional: limpiar/trimming
        if isinstance(contenido, str):
            contenido = contenido.strip()

        # usar transaction por seguridad
        with transaction.atomic():
            cot.acu_e = contenido
            cot.save(update_fields=["acu_e"])

        return Response({
            "status": "ok",
            "msg": "Condiciones guardadas correctamente.",
            "numero": cot.numero,
            "condiciones": cot.acu_e or ""
        }, status=200)

    except Exception as e:
        import traceback
        print("Error en condiciones_generales:", traceback.format_exc())
        return Response({"error": str(e)}, status=500)

@csrf_exempt
@api_view(["GET", "POST"])
def generar_codigo_cotizacion(request, num_reg):

    try:
        # =========================
        # OBTENER COTIZACIÓN
        # =========================
        cot = DashboardCotizacion.objects.get(num_reg=num_reg)

        # Si ya tiene código, no inventamos nada
        if cot.numero:
            return JsonResponse({
                "success": True,
                "codigo": cot.numero,
                "exists": True
            })

        # =========================
        # AÑO
        # =========================
        year_full = cot.anno              # ej: 2025
        year = year_full[-2:]             # ej: 25

        # =========================
        # ÁREA
        # =========================
        area = str(cot.area_codigo)

        # =========================
        # ÚLTIMA SECUENCIA POR AÑO + ÁREA
        # =========================
        qs = (
            DashboardCotizacion.objects
            .filter(
                anno=year_full,
                area_codigo=area,
                numero__isnull=False,
            )
            .exclude(numero="")
            .values_list("numero", flat=True)
        )

        max_corr = 0
        len_base = len(year) + len(area)  # 2 + 1 = 3

        for num in qs:
            try:
                corr = int(num[len_base:len_base+3])   # los 3 dígitos del correlativo
                max_corr = max(max_corr, corr)
            except Exception:
                continue

        correlativo = max_corr + 1
        correlativo_str = str(correlativo).zfill(3)

        # =========================
        # BASE DEL CÓDIGO
        # =========================
        base_codigo = f"{year}{area}{correlativo_str}"

        # =========================
        # VERSIÓN (A, B, C...)
        # =========================
        existentes = (
            DashboardCotizacion.objects
            .filter(numero__startswith=base_codigo)
            .values_list("numero", flat=True)
        )

        if not existentes:
            version = "A"
        else:
            letras = [c[len(base_codigo)] for c in existentes]
            version = ascii_uppercase[ascii_uppercase.index(max(letras)) + 1]

        # =========================
        # CLIENTE
        # =========================
        cliente = vc_tab_clientes.objects.get(codigo=cot.cliente_codigo)
        iniciales = cliente.iniciales

        # =========================
        # TIPO
        # =========================
        tipo = cot.cotit

        # =========================
        # CÓDIGO FINAL
        # =========================
        codigo_final = f"{base_codigo}{version}-{iniciales}-{tipo}"

        # =========================
        # SOLO GUARDA EN POST
        # =========================
        if request.method == "POST":
            cot.numero = codigo_final
            cot.save(update_fields=["numero"])

        return JsonResponse({
            "success": True,
            "codigo": codigo_final,
            "preview": request.method == "GET"
        })

    except DashboardCotizacion.DoesNotExist:
        return JsonResponse({
            "success": False,
            "error": "Cotización no encontrada"
        }, status=404)

@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def descuento_cotizacion(request, num_reg):

    num_reg = str(num_reg)

    cot = DashboardCotizacion.objects.filter(num_reg=num_reg).first()

    if not cot:
        return Response({}, status=404)

    # ============================
    # 🔹 GET → OBTENER
    # ============================
    if request.method == "GET":

        afecto_map = {
            "T": "t",
            "S": "su",
            "M": "ser",
        }

        return Response({
            "aplicar": bool(cot.des_a),
            "afecto": afecto_map.get(cot.des_t, "t"),
            "porcentaje": str(cot.des_p or ""),
            "importe": str(cot.des_m or ""),
        })

    # ============================
    # 🔹 POST → GUARDAR + RECALCULAR TOTAL
    # ============================

    data = request.data

    afecto_reverse = {
        "t": "T",
        "su": "S",
        "ser": "M",
    }

    # ----------------------------
    # 📌 Guardar descuento
    # ----------------------------

    aplicar = bool(data.get("aplicar"))

    afecto_front = data.get("afecto", "t")

    cot.des_a = 1 if aplicar else 0
    cot.des_t = afecto_reverse.get(afecto_front, "T")

    cot.des_p = data.get("porcentaje") or None
    cot.des_m = data.get("importe") or None

    # ----------------------------
    # 🔁 RECALCULAR TOTALES
    # ----------------------------

    total_suministros = CotiSuministros.objects.filter(
        num_reg=num_reg,
        nig=0
    ).aggregate(
        total=Coalesce(
            Sum(
                ExpressionWrapper(
                    F("tot") * F("can"),
                    output_field=DecimalField(max_digits=18, decimal_places=2),
                )
            ),
            Decimal("0.00")
        )
    )["total"]

    total_servicios = CotiServicios.objects.filter(
        num_reg=num_reg,
        nig=0
    ).aggregate(
        total=Coalesce(
            Sum("tot"),
            Decimal("0.00")
        )
    )["total"]

    total_general = total_suministros + total_servicios

    # ----------------------------
    # 🔻 Aplicar descuento
    # ----------------------------

    importe_desc = Decimal(str(data.get("importe") or 0))

    if aplicar and importe_desc > 0:

        if afecto_front == "t":
            total_general -= importe_desc

        elif afecto_front == "su":
            total_general = (total_suministros - importe_desc) + total_servicios

        elif afecto_front == "ser":
            total_general = total_suministros + (total_servicios - importe_desc)

    if total_general < 0:
        total_general = Decimal("0.00")

    cot.tot_c = total_general

    # ----------------------------
    # 💾 Guardar todo
    # ----------------------------

    cot.save(update_fields=[
        "tot_c",
        "des_a",
        "des_t",
        "des_p",
        "des_m",
    ])

    return Response({"ok": True})

def build_cotizacion_pdf_context(num_reg):

    # =========================
    # CABECERA (solo campos usados)
    # =========================
    cotizacion = (
        DashboardCotizacion.objects
        .only(
            "numero", "fecha", "referencia", "cliente_codigo",
            "nombr", "cargr", "teler", "movir", "mailr",
            "nombc", "telec", "mov1c", "mov2c", "mov3c", "mailc",
            "nombt", "telet", "mov1t", "mov2t", "mov3t", "mailt",
            "plazo", "tot_d", "por_c", "tot_s", "tcamb",
            "fpago", "lugar",
            "tmone", "igv",
            "valid", "acu_s",
            "tot_c", "acu_e", "des_m", "num_reg",
        )
        .filter(num_reg=num_reg)
        .first()
    )

    if not cotizacion:
        return None

    # =========================
    # DETALLES
    # =========================
    suministros_qs = (
        CotiSuministros.objects
        .filter(num_reg=num_reg)
        .order_by("cog", "nig", "num")
        .iterator()
    )

    servicios_qs = (
        CotiServicios.objects
        .filter(num_reg=num_reg)
        .order_by("cog", "nig", "num")
        .iterator()
    )

    # =========================
    # CABECERA CONTEXT
    # =========================
    cabecera = {
        "numero": cotizacion.numero,
        "num_reg": cotizacion.num_reg,
        "fecha": cotizacion.fecha,
        "referencia": cotizacion.referencia,
        "cliente": cotizacion.cliente_codigo,
        "atencion": {
            "nombre": cotizacion.nombr,
            "cargo": cotizacion.cargr,
            "telefono": cotizacion.teler or cotizacion.movir,
            "correo": cotizacion.mailr,
        },
        "comercial": {
            "nombre": cotizacion.nombc,
            "telefono": cotizacion.telec,
            "movil1": cotizacion.mov1c,
            "movil2": cotizacion.mov2c,
            "movil3": cotizacion.mov3c,
            "correo": cotizacion.mailc,
        },
        "tecnico": {
            "nombre": cotizacion.nombt,
            "telefono": cotizacion.telet,
            "movil1": cotizacion.mov1t,
            "movil2": cotizacion.mov2t,
            "movil3": cotizacion.mov3t,
            "correo": cotizacion.mailt,
        },
        "tiempo_entrega": {
            "suministros": {
                "cantidad": cotizacion.plazo,
                "tipo": "Días" if cotizacion.tot_d == "D" else "Semanas" if cotizacion.tot_d == "S" else "Meses",
            },
            "servicios": {
                "cantidad": cotizacion.por_c,
                "tipo": "Días" if cotizacion.tot_s == "D" else "Semanas" if cotizacion.tot_s == "S" else "Meses",
            },
        },
        "forma_pago": cotizacion.fpago,
        "lugar_entrega": cotizacion.lugar,
        "moneda": "Dólares" if cotizacion.tmone == "D" else "Soles",
        "moneda_simbolo": "USD" if cotizacion.tmone == "D" else "PEN",
        "incluye_igv": cotizacion.igv == "S",
        "validez": {
            "cantidad": cotizacion.valid,
            "tipo": "Días" if cotizacion.acu_s == "D" else "Semanas" if cotizacion.acu_s == "S" else "Meses",
        },
    }

    # =========================
    # CONVERSIÓN MONEDA
    # =========================
    def convertir(valor, cotizacion):
        if cotizacion.tmone == "S" and cotizacion.tcamb:
            return (valor or Decimal("0.00")) * cotizacion.tcamb
        return valor or Decimal("0.00")

    # =========================
    # SUMINISTROS
    # =========================
    grupos = OrderedDict()

    for s in suministros_qs:

        if not s.cog:
            continue

        if s.nig == 0:

            can = s.can or Decimal("1.00")
            tot = convertir(s.tot, cotizacion)

            grupos[s.cog] = {
                "cog": s.cog,
                "titulo": s.nog,
                "mov": s.mov,
                "entrega":s.ent or 0,
                "cantidad": can,
                "total": tot,
                "total_grupo": tot * can,
                "subtotal_pu_items": Decimal("0.00"),
                "subtotal_tot_items": Decimal("0.00"),
                "items": [],
            }

        elif s.nig == 1 and s.cog in grupos:

            pu = convertir(s.puc, cotizacion)
            tot = convertir(s.tot, cotizacion)

            grupos[s.cog]["items"].append({
                "codigo": s.cod,
                "descripcion": s.des,
                "unidad": s.tde,
                "entrega":s.ent or 0,
                "cantidad": s.can or 0,
                "precio_unitario": pu,
                "total": tot,
            })

            grupos[s.cog]["subtotal_pu_items"] += pu
            grupos[s.cog]["subtotal_tot_items"] += tot

    suministros = list(grupos.values())
    total_suministros = sum(
        (g["total_grupo"] for g in suministros),
        Decimal("0.00"),
    )

    # =========================
    # SERVICIOS
    # =========================
    servicios_grupos = OrderedDict()

    for s in servicios_qs:

        if not s.cog:
            continue

        if s.nig == 0:

            can = s.can or Decimal("1.00")
            tot = convertir(s.tot, cotizacion)

            servicios_grupos[s.cog] = {
                "cog": s.cog,
                "titulo": s.nog,
                "mov": s.mov,
                "cantidad": can,
                "total": tot,
                "total_servicio": tot * can,
                "detalle": s.tog or "",
                "subtotal_pu_items": Decimal("0.00"),
                "subtotal_tot_items": Decimal("0.00"),
                "items": [],
            }

        elif s.nig == 1 and s.cog in servicios_grupos:

            tot = convertir(s.tot, cotizacion)
            pu = convertir(s.puc, cotizacion)

            servicios_grupos[s.cog]["items"].append({
                "codigo": s.cod,
                "descripcion": s.des,
                "proveedor": s.pro,
                "unidad": s.tde,
                "cantidad": s.can or Decimal("0.00"),
                "precio_unitario": pu,
                "total": tot,
            })

            servicios_grupos[s.cog]["subtotal_pu_items"] += pu
            servicios_grupos[s.cog]["subtotal_tot_items"] += tot

    servicios = list(servicios_grupos.values())

    total_servicios = sum(
        (s["total_servicio"] for s in servicios),
        Decimal("0.00"),
    )

    # =========================
    # CONTEXT FINAL
    # =========================
    descuento = convertir(cotizacion.des_m, cotizacion)
    total_bruto = convertir(cotizacion.tot_c, cotizacion)

    total_final = total_bruto - descuento

    # =========================
    # ÍNDICE DINÁMICO
    # =========================
    indice = []
    contador = 1

    # 1
    indice.append({
        "numero": contador,
        "titulo": "Presupuesto General"
    })
    contador += 1

    # SUMINISTROS
    for g in suministros:
        indice.append({
            "numero": contador,
            "titulo": f"Detalle: {g['titulo']}"
        })
        contador += 1

    # SERVICIOS
    for s in servicios:
        indice.append({
            "numero": contador,
            "titulo": f"Detalle: {s['titulo']}"
        })
        contador += 1

    # CONDICIONES
    indice.append({
        "numero": contador,
        "titulo": "Condiciones Generales y Garantías"
    })

    # =========================
    # NUMERACIÓN DE SECCIONES
    # =========================
    secciones = {
        "presupuesto": 1,
    }

    contador = 2

    # Suministros
    for g in suministros:
        g["seccion"] = contador
        contador += 1

    # Servicios
    for s in servicios:
        s["seccion"] = contador
        contador += 1

    # Condiciones
    secciones["condiciones"] = contador

    return {
        "cabecera": cabecera,
        "suministros": suministros,
        "servicios": servicios,
        "totales": {
            "suministros": total_suministros,
            "servicios": total_servicios,
            "descuento": descuento,
            "total_cotizacion": total_final,
            "moneda": cabecera["moneda"],
            "moneda_simbolo": cabecera["moneda_simbolo"],
            "incluye_igv": cabecera["incluye_igv"],
        },
        "condiciones_generales": {
            "condiciones": cotizacion.acu_e,
        },
        "indice": indice,
        "secciones": secciones,
    }

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def cotizacion_pdf_context(request, num_reg):
    context = build_cotizacion_pdf_context(num_reg)

    if not context:
        return Response({"error": "Cotización no existe"}, status=404)

    return Response(context)

@csrf_exempt
def cotizacion_pdf_preview(request, num_reg):
    context = build_cotizacion_pdf_context(num_reg)

    if not context:
        return HttpResponse("Cotización no existe", status=404)

    return render(request, "reportes/cotizacion_pdf.html", context)

@csrf_exempt
def cotizacion_reporte_html(request, num_reg):
    context = build_cotizacion_pdf_context(num_reg)

    if not context:
        return HttpResponse("Cotización no existe", status=404)

    return render(request, "reportes/cotizacion_pdf.html", context)

def descargar_cotizacion_word(request, num_reg):
    # Asumimos que build_cotizacion_pdf_context ya trae toda la data necesaria
    context = build_cotizacion_pdf_context(num_reg)
    if not context:
        return HttpResponse("La cotización no existe", status=404)

    # --- PALETA DE COLORES CEBRA ---
    COLOR_GRIS_CLARO = "#FFFFFF"  # Fila A (Grisáceo)
    COLOR_AZUL_SUAVE = "#FEFEFF"  # Fila B (Azulado)
    # ------------------------------

    template_path = os.path.join(settings.BASE_DIR, 'cotizaciones_api', 'templates', 'reportes', 'plantilla_word.docx')
    
    try:
        doc = DocxTemplate(template_path)
        moneda = context['totales'].get('moneda', 'Dólares')

        def formatear_moneda(valor):
            simbolo = "$" if moneda == "Dólares" else "S/."
            return f"{simbolo} {valor:,.2f}"
        
        # 0. Inicializamos contador global para alternancia de colores
        fila_idx = 0

        # 1. Procesamiento de suministros
        for g in context.get('suministros', []):
            # Asignación de color para el sombreado de la fila
            g['bg_color'] = COLOR_AZUL_SUAVE if fila_idx % 2 == 0 else COLOR_GRIS_CLARO
            fila_idx += 1
            
            g['total_g_f'] = formatear_moneda(g.get('total_grupo', 0))
            g['subtotal_f'] = formatear_moneda(g.get('subtotal_tot_items', 0))
            g['unitario_f'] = formatear_moneda(g.get('total', 0))
            g['cant_f'] = f"{g.get('cantidad', 0):,.2f}"
            
            g['filas'] = g.get('items', [])
            for item in g['filas']:
                desc = item.get('descripcion', '') or ''
                # Limpieza de HTML básico para descripciones de items
                item['desc_f'] = RichText(desc.replace('<br>', '\n').replace('<br/>', '\n'))
                item['precio_f'] = formatear_moneda(item.get('precio_unitario', 0))
                item['total_f'] = formatear_moneda(item.get('total', 0))
                item['cant_f'] = f"{item.get('cantidad', 0):,.2f}"

        # 2. Procesamiento de servicios
        for s in context.get('servicios', []):
            # Continuamos la alternancia basándonos en el contador global
            s['bg_color'] = COLOR_AZUL_SUAVE if fila_idx % 2 == 0 else COLOR_GRIS_CLARO
            fila_idx += 1
            
            s['total_g_f'] = formatear_moneda(s.get('total_servicio', 0))
            s['unitario_f'] = formatear_moneda(s.get('total', 0))
            s['cant_f'] = f"{s.get('cantidad', 0):,.2f}"
            
            detalle_raw = s.get('detalle', '') or ''
            rt = RichText()
            
            # Decodificación y limpieza de bloques HTML
            texto = unescape(detalle_raw).replace('&nbsp;', ' ')
            bloques = re.split(r'(<li>|<p>|<ul>|</ul>|</p>|</li>)', texto)
            
            dentro_de_lista = False
            for i, parte in enumerate(bloques):
                if '<ul>' in parte:
                    dentro_de_lista = True
                    continue
                if '</ul>' in parte:
                    dentro_de_lista = False
                    continue
                    
                contenido = re.sub(r'<[^>]+>', '', parte).strip()
                if not contenido:
                    continue
                    
                # Formateo de títulos (negrita) y viñetas
                if contenido.endswith(':'):
                    if len(rt.xml) > 0: rt.add('\n')
                    rt.add(contenido, bold=True)
                    rt.add('\n')
                elif '<li>' in bloques[i-1]:
                    rt.add(f" • {contenido}\n")
                elif '<p>' in bloques[i-1] or not dentro_de_lista:
                    rt.add(f"{contenido}\n")

            s['detalle_f'] = rt

        # 2.5 Procesamiento de Condiciones Generales (Estilo para Cuadro Dinámico)
        cond_raw = context.get('condiciones_generales', {}).get('condiciones', '') or ''
        rt_cond = RichText()

        COLOR_PRO = "444444"
        TAMANO_PRO = 18  # 9pt

        if cond_raw:
            texto = unescape(cond_raw).replace('&nbsp;', ' ').replace('\xa0', ' ')
            bloques = re.split(r'(<li>|<p>|<ul>|</ul>|</p>|</li>)', texto)

            dentro_de_lista = False

            for i, parte in enumerate(bloques):
                if '<ul>' in parte:
                    dentro_de_lista = True
                    continue
                if '</ul>' in parte:
                    dentro_de_lista = False
                    continue

                contenido = re.sub(r'<[^>]+>', '', parte).strip()
                if not contenido:
                    continue

                es_titulo = contenido.endswith(':') or (contenido.isupper() and len(contenido) > 3)

                if es_titulo:
                    if len(rt_cond.xml) > 0:
                        rt_cond.add('\n')
                    rt_cond.add(contenido, bold=True, color=COLOR_PRO, size=TAMANO_PRO)
                    rt_cond.add('\n')

                elif i > 0 and '<li>' in bloques[i-1]:
                    rt_cond.add(f" • {contenido}\n", color=COLOR_PRO, size=TAMANO_PRO)

                else:
                    rt_cond.add(f"{contenido}\n", color=COLOR_PRO, size=TAMANO_PRO)

        context['condiciones_generales']['texto_f'] = rt_cond

        # 3. Procesamiento de totales finales
        t = context['totales']
        t['desc_f'] = formatear_moneda(t.get('descuento', 0))
        t['total_f'] = formatear_moneda(t.get('total_cotizacion', 0))
        
        # Renderizado único del documento
        doc.render(context)

        # Preparación de la respuesta de descarga
        buffer = io.BytesIO()
        doc.save(buffer)
        content = buffer.getvalue()
        buffer.close()

        response = HttpResponse(
            content,
            content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )
        response['Content-Disposition'] = f'attachment; filename="Cotizacion_{num_reg}.docx"'
        return response

    except Exception as e:
        return HttpResponse(f"Error técnico en el servidor: {str(e)}", status=500) 

def cotizacion_pdf(request, num_reg):
    context = build_cotizacion_pdf_context(num_reg)

    if not context:
        return HttpResponse("Cotización no existe", status=404)

    # 📌 Definimos la carpeta base de assets para evitar repetir código
    assets_dir = Path(settings.BASE_DIR) / "frontend" / "src" / "assets"

    # 📌 Rutas de Imágenes Principales
    context["logo_path"] = (assets_dir / "logo.png").as_uri()
    context["header_path"] = (assets_dir / "encabezado-reporte.png").as_uri()

    # 📌 Rutas del Nuevo Pie de Página (Basado en tus archivos)
    context["footer_bg_path"] = (assets_dir / "pie pagina-reporte.png").as_uri()
    context["sgs_path"] = (assets_dir / "sgs.png").as_uri()
    context["homologada_path"] = (assets_dir / "empresa homologada.png").as_uri()
    context["mega_path"] = (assets_dir / "mega.png").as_uri()
    context["correo_path"] = (assets_dir / "correo cormercial.png").as_uri()

    html_string = render_to_string("reportes/cotizacion_pdf.html", context)

    # Generación del PDF
    html = HTML(
        string=html_string, 
        base_url=settings.BASE_DIR.as_uri()
    )

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="cotizacion_{num_reg}.pdf"'

    # Nota: Usamos optimización de imágenes para evitar que el PDF pese demasiado
    html.write_pdf(response)
    return response

@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def eliminar_cotizacion(request, num_reg):
    """
    Elimina una cotización y todos sus registros relacionados.
    """
    try:
        with transaction.atomic():

            # 1. Verificar existencia de la cotización
            cotizacion = DashboardCotizacion.objects.filter(
                num_reg=num_reg
            ).first()

            if not cotizacion:
                return Response(
                    {"error": "La cotización no existe"},
                    status=404
                )

            # 2. Eliminar dependencias
            CotiSuministros.objects.filter(num_reg=num_reg).delete()
            CotiServicios.objects.filter(num_reg=num_reg).delete()
            CotiMensajes.objects.filter(num_reg=num_reg).delete()
            CotiSeguimiento.objects.filter(num_reg=num_reg).delete()

            # 3. Eliminar cotización principal
            cotizacion.delete()

        return Response(
            {"message": "Cotización eliminada correctamente"},
            status=200
        )

    except Exception as e:
        return Response(
            {"error": str(e)},
            status=500
        )

@api_view(["PATCH"])
@permission_classes([IsAuthenticated])
def enviar_cotizacion_aprobacion(request, num_reg):
    """
    Cambia el estado de envio de la cotización:
    0 -> 1
    1 -> 2
    """
    try:
        with transaction.atomic():
            cotizacion = DashboardCotizacion.objects.filter(num_reg=num_reg).first()

            if not cotizacion:
                return Response({"error": "La cotización no existe"}, status=404)

            # Cambiar valor de envio según el caso
            if cotizacion.envio == 0:
                cotizacion.envio = 1
            elif cotizacion.envio == 1:
                cotizacion.envio = 2
            else:
                return Response({"error": f"Estado de envio inválido: {cotizacion.envio}"}, status=400)

            cotizacion.save()

        return Response({"message": "Cotización enviada a aprobación correctamente"}, status=200)

    except Exception as e:
        return Response({"error": str(e)}, status=500)

@api_view(["PATCH"])
@permission_classes([IsAuthenticated])
def cerrar_cotizacion(request, num_reg):
    """
    Cierra la cotización:
    envio: 2 -> 3
    """
    try:
        with transaction.atomic():
            cotizacion = DashboardCotizacion.objects.filter(num_reg=num_reg).first()

            if not cotizacion:
                return Response(
                    {"error": "La cotización no existe"},
                    status=404
                )

            if cotizacion.envio != 2:
                return Response(
                    {
                        "error": "La cotización no está lista para cerrarse",
                        "estado_actual": cotizacion.envio
                    },
                    status=400
                )

            cotizacion.envio = 3
            cotizacion.save(update_fields=["envio"])

        return Response(
            {"message": "Cotización cerrada correctamente"},
            status=200
        )

    except Exception as e:
        return Response(
            {"error": str(e)},
            status=500
        )

@csrf_exempt
@api_view(["PATCH"])
@permission_classes([IsAuthenticated])
def asignar_regus(request, num_reg):
    """
    Actualiza los campos 'regus' y 'referencia' de una cotización según num_reg.
    """
    try:
        cotizacion = DashboardCotizacion.objects.get(num_reg=num_reg)
    except DashboardCotizacion.DoesNotExist:
        return Response(
            {"detail": "Cotización no encontrada"},
            status=404
        )

    # Obtener datos del request
    regus = request.data.get("regus")
    referencia = request.data.get("referencia")

    if not regus and not referencia:
        return Response(
            {"detail": "Debe enviar al menos 'regus' o 'referencia' para actualizar"},
            status=400
        )

    campos_a_actualizar = []

    if regus:
        cotizacion.regus = regus
        campos_a_actualizar.append("regus")

    if referencia:
        cotizacion.referencia = referencia
        campos_a_actualizar.append("referencia")

    cotizacion.save(update_fields=campos_a_actualizar)

    return Response(
        {
            "message": "Cotización actualizada correctamente",
            "num_reg": cotizacion.num_reg,
            "regus": cotizacion.regus,
            "referencia": cotizacion.referencia,
        },
        status=200
    )

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def crear_nueva_version_cotizacion(request, num_reg):
    try:
        with transaction.atomic():

            print("🔹 Buscando cotización base")
            base = DashboardCotizacion.objects.filter(num_reg=num_reg).first()
            if not base:
                return Response({"error": "La cotización no existe"}, status=404)

            # 🔒 SOLO SI TIENE COTIN
            if not base.numero:
                return Response(
                    {
                        "error": "La cotización aún no tiene COTIN. No se puede generar una nueva versión."
                    },
                    status=400,
                )

            print("🔹 Generando nuevo cotin")
            nuevo_cotin = siguiente_version(base.numero)

            print("🔹 Creando nueva cotización")
            
            # Extraemos el código del usuario desde el token (tal como en tu endpoint)
            from rest_framework_simplejwt.authentication import JWTAuthentication
            jwt_auth = JWTAuthentication()
            header = jwt_auth.get_header(request)
            raw_token = jwt_auth.get_raw_token(header)
            validated_token = jwt_auth.get_validated_token(raw_token)
            usuario_codigo = validated_token.get("user_id") # Este es tu usuario_usu

            nueva_cotizacion = deepcopy(base)
            nueva_cotizacion.pk = None
            nueva_cotizacion.num_reg = None
            nueva_cotizacion.numero = nuevo_cotin
            nueva_cotizacion.estado_codigo = 2
            nueva_cotizacion.envio = 0
            
            # Asignamos el código del usuario que realiza la acción
            nueva_cotizacion.regus = usuario_codigo 
            
            nueva_cotizacion.save()

            # =====================================================
            # 🔹 SUMINISTROS (GRUPOS + ÍTEMS)
            # =====================================================
            print("🔹 Replicando SUMINISTROS")

            suministros_origen = list(
                CotiSuministros.objects.filter(num_reg=num_reg).order_by("num")
            )

            ultimo_num = (
                CotiSuministros.objects.aggregate(max_num=Max("num"))["max_num"] or 0
            )

            nuevos = []
            contador = ultimo_num + 1

            for s in suministros_origen:
                nuevo = deepcopy(s)
                nuevo.pk = None
                nuevo.num = contador
                nuevo.num_reg = nueva_cotizacion.num_reg
                nuevos.append(nuevo)
                contador += 1

            CotiSuministros.objects.bulk_create(nuevos)

            # =====================================================
            # 🔹 SERVICIOS (GRUPOS + ÍTEMS)
            # =====================================================
            print("🔹 Replicando SERVICIOS")

            servicios_origen = list(
                CotiServicios.objects.filter(num_reg=num_reg).order_by("num")
            )

            ultimo_num_serv = (
                CotiServicios.objects.aggregate(max_num=Max("num"))["max_num"] or 0
            )

            nuevos_serv = []
            contador = ultimo_num_serv + 1

            for s in servicios_origen:
                nuevo = deepcopy(s)
                nuevo.pk = None
                nuevo.num = contador
                nuevo.num_reg = nueva_cotizacion.num_reg
                nuevos_serv.append(nuevo)
                contador += 1

            CotiServicios.objects.bulk_create(nuevos_serv)

            # =====================================================
            # 🔹 MENSAJES
            # =====================================================
            print("🔹 Replicando MENSAJES")

            mensajes_origen = list(
                CotiMensajes.objects.filter(num_reg=num_reg).order_by("dat")
            )

            offset = 0
            nuevos_mensajes = []

            for m in mensajes_origen:
                nuevo = deepcopy(m)
                nuevo.pk = None
                nuevo.dat = now() + timedelta(milliseconds=offset)
                nuevo.num_reg = nueva_cotizacion.num_reg
                nuevos_mensajes.append(nuevo)
                offset += 1

            CotiMensajes.objects.bulk_create(nuevos_mensajes)

            # =====================================================
            # 🔹 SEGUIMIENTO
            # =====================================================
            print("🔹 Replicando SEGUIMIENTO")

            seguimiento_origen = list(
                CotiSeguimiento.objects.filter(num_reg=num_reg).order_by("dat")
            )

            offset = 0
            nuevos_seg = []

            for seg in seguimiento_origen:
                nuevo = deepcopy(seg)
                nuevo.pk = None
                nuevo.dat = now() + timedelta(milliseconds=offset)
                nuevo.num_reg = nueva_cotizacion.num_reg
                nuevo.num = seg.num
                nuevos_seg.append(nuevo)
                offset += 1

            CotiSeguimiento.objects.bulk_create(nuevos_seg)

        return Response(
            {
                "message": "Nueva versión creada correctamente",
                "num_reg": nueva_cotizacion.num_reg,
                "cotin": nueva_cotizacion.numero,
                "num_reg_origen": base.num_reg,  # 👈 útil para invalidaciones finas
            },
            status=201,
        )

    except Exception as e:
        print("❌ ERROR REAL:", e)
        raise e

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def generar_copiar_cotizacion(request, num_reg):
    """
    Crea una copia de la cotización indicada sin asignarle COTIN (numero).
    La copia se genera con estado pendiente de envío (estado_codigo=2).
    """
    try:
        with transaction.atomic():

            print("🔹 Buscando cotización base")
            base = DashboardCotizacion.objects.filter(num_reg=num_reg).first()
            if not base:
                return Response({"error": "La cotización no existe"}, status=404)

            print("🔹 Creando copia sin COTIN")
            
            # Extraemos el código del usuario desde el token
            from rest_framework_simplejwt.authentication import JWTAuthentication
            jwt_auth = JWTAuthentication()
            header = jwt_auth.get_header(request)
            raw_token = jwt_auth.get_raw_token(header)
            validated_token = jwt_auth.get_validated_token(raw_token)
            usuario_codigo = validated_token.get("user_id")

            nueva_cotizacion = deepcopy(base)
            nueva_cotizacion.pk = None
            nueva_cotizacion.num_reg = None
            nueva_cotizacion.numero = None
            nueva_cotizacion.estado_codigo = 2
            nueva_cotizacion.envio = 0
            
            # Marcamos quién registró la copia
            nueva_cotizacion.regus = usuario_codigo

            # =========================
            # 🔹 REFERENCIA (MARCAR COPIA)
            # =========================
            if base.referencia:
                nueva_cotizacion.referencia = f"{base.referencia} - COPIA"
            else:
                nueva_cotizacion.referencia = f"COPIA DE COTIZACIÓN {num_reg}"

            nueva_cotizacion.save()

            # =====================================================
            # 🔹 SUMINISTROS (GRUPOS + ÍTEMS)
            # =====================================================
            print("🔹 Replicando SUMINISTROS")

            suministros_origen = list(
                CotiSuministros.objects.filter(num_reg=num_reg).order_by("num")
            )

            ultimo_num = (
                CotiSuministros.objects.aggregate(max_num=Max("num"))["max_num"] or 0
            )

            nuevos = []
            contador = ultimo_num + 1

            for s in suministros_origen:
                nuevo = deepcopy(s)
                nuevo.pk = None
                nuevo.num = contador
                nuevo.num_reg = nueva_cotizacion.num_reg
                nuevos.append(nuevo)
                contador += 1

            CotiSuministros.objects.bulk_create(nuevos)

            # =====================================================
            # 🔹 SERVICIOS (GRUPOS + ÍTEMS)
            # =====================================================
            print("🔹 Replicando SERVICIOS")

            servicios_origen = list(
                CotiServicios.objects.filter(num_reg=num_reg).order_by("num")
            )

            ultimo_num_serv = (
                CotiServicios.objects.aggregate(max_num=Max("num"))["max_num"] or 0
            )

            nuevos_serv = []
            contador = ultimo_num_serv + 1

            for s in servicios_origen:
                nuevo = deepcopy(s)
                nuevo.pk = None
                nuevo.num = contador
                nuevo.num_reg = nueva_cotizacion.num_reg
                nuevos_serv.append(nuevo)
                contador += 1

            CotiServicios.objects.bulk_create(nuevos_serv)

            # =====================================================
            # 🔹 MENSAJES
            # =====================================================
            print("🔹 Replicando MENSAJES")

            mensajes_origen = list(
                CotiMensajes.objects.filter(num_reg=num_reg).order_by("dat")
            )

            offset = 0
            nuevos_mensajes = []

            for m in mensajes_origen:
                nuevo = deepcopy(m)
                nuevo.pk = None
                nuevo.dat = now() + timedelta(milliseconds=offset)
                nuevo.num_reg = nueva_cotizacion.num_reg
                nuevos_mensajes.append(nuevo)
                offset += 1

            CotiMensajes.objects.bulk_create(nuevos_mensajes)

            # =====================================================
            # 🔹 SEGUIMIENTO
            # =====================================================
            print("🔹 Replicando SEGUIMIENTO")

            seguimiento_origen = list(
                CotiSeguimiento.objects.filter(num_reg=num_reg).order_by("dat")
            )

            offset = 0
            nuevos_seg = []

            for seg in seguimiento_origen:
                nuevo = deepcopy(seg)
                nuevo.pk = None
                nuevo.dat = now() + timedelta(milliseconds=offset)
                nuevo.num_reg = nueva_cotizacion.num_reg
                nuevo.num = seg.num
                nuevos_seg.append(nuevo)
                offset += 1

            CotiSeguimiento.objects.bulk_create(nuevos_seg)

        return Response(
            {
                "message": "Copia de cotización creada correctamente sin COTIN",
                "num_reg": nueva_cotizacion.num_reg,
                "cotin": nueva_cotizacion.numero,  # será None
            },
            status=201,
        )

    except Exception as e:
        print("❌ ERROR REAL:", e)
        raise e

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def generar_codigo_view(request, numero):
    """
    Retorna el código (cotin) asociado a la cotización.
    Si la cotización no existe → 404
    """
    try:
        cot = DashboardCotizacion.objects.filter(numero=numero).first()
        if not cot:
            return Response(
                {"error": f"No se encontró la cotización #{numero}"},
                status=404
            )

        return Response({
            "numero": cot.numero,
            "codigo": cot.numero  # Es lo mismo que cotin
        })

    except Exception as e:
        import traceback
        print("Error en generar_codigo_view:", traceback.format_exc())
        return Response({"error": str(e)}, status=500)

@api_view(["PATCH"])
@permission_classes([IsAuthenticated])
def cambiar_estado_cotizacion(request, num_reg):
    """
    Cambia el estado de una cotización al estado indicado.
    Ejemplo: 1 → 3, 3 → 2, etc.
    """
    try:
        estado_codigo = request.data.get("estado_codigo")

        if estado_codigo is None:
            return Response(
                {"error": "Debe enviar estado_codigo"},
                status=400
            )

        with transaction.atomic():
            cotizacion = DashboardCotizacion.objects.filter(num_reg=num_reg).first()

            if not cotizacion:
                return Response(
                    {"error": "La cotización no existe"},
                    status=404
                )

            # Validar que el estado exista y esté activo
            estado = vc_tab_estado.objects.filter(
                codigo=estado_codigo,
                activo=True
            ).first()

            if not estado:
                return Response(
                    {"error": "Estado inválido o inactivo"},
                    status=400
                )

            cotizacion.estado_codigo = estado.codigo
            cotizacion.save()

        return Response(
            {"message": "Estado de la cotización actualizado correctamente"},
            status=200
        )

    except Exception as e:
        return Response(
            {"error": str(e)},
            status=500
        )

@api_view(["PATCH"])
@permission_classes([IsAuthenticated])
def retornar_cotizacion(request, num_reg):
    """
    Retorna una cotización a estado editable.
    Acción: envio = 0
    """

    try:
        with transaction.atomic():
            cotizacion = DashboardCotizacion.objects.filter(num_reg=num_reg).first()

            if not cotizacion:
                return Response(
                    {"error": "La cotización no existe"},
                    status=404
                )

            # 🔒 Opcional: validar que esté enviada
            if cotizacion.envio == 0:
                return Response(
                    {"message": "La cotización ya está en estado editable"},
                    status=200
                )

            cotizacion.envio = 0
            cotizacion.save(update_fields=["envio"])

        return Response(
            {"message": "La cotización fue retornada correctamente"},
            status=200
        )

    except Exception as e:
        return Response(
            {"error": str(e)},
            status=500
        )

#========================================================================================

##===================##
## OBJETIVOS ANUALES ##
##===================##
@api_view(["GET", "POST", "PUT"])
def objetivos_anuales(request):
    # Mantenemos la referencia pero ya no será el filtro principal
    usuario_codigo = request.user.usuario_usu

    # =================
    # LISTAR (Global)
    # =================
    if request.method == "GET":
        anno = request.query_params.get("anno")

        if not anno:
            return Response({"error": "Debe enviar el año"}, status=status.HTTP_400_BAD_REQUEST)

        # CAMBIO CLAVE: Cambiamos .get() por .filter() para traer todos los objetivos del año.
        # Quitamos: encargado=usuario_codigo
        objetivos = ObjetivoAnual.objects.prefetch_related("areas").filter(
            anno=anno,
            activo=True
        )

        if objetivos.exists():
            # many=True porque ahora sumaremos todas las metas activas del año
            serializer = ObjetivoAnualSerializer(objetivos, many=True)
            return Response(serializer.data)
        else:
            return Response(
                {"message": "No existen objetivos activos para ese año"},
                status=status.HTTP_200_OK
            )

    # =================
    # CREAR (Global)
    # =================
    if request.method == "POST":
        # Ahora desactivamos TODOS los objetivos del año indicado para "resetear" la meta global
        anno_post = request.data.get("anno")
        ObjetivoAnual.objects.filter(
            anno=anno_post,
            activo=True
        ).update(activo=False)

        serializer = ObjetivoAnualSerializer(data=request.data)
        if serializer.is_valid():
            # El encargado sigue siendo quien CREA el registro por auditoría
            serializer.save(encargado=usuario_codigo) 
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    # =================
    # ACTUALIZAR
    # =================
    if request.method == "PUT":
        anno = request.data.get("anno")
        # Aquí podrías usar el ID del objetivo para ser más preciso, 
        # pero si mantienes anno, filtramos el que esté activo.
        try:
            objetivo = ObjetivoAnual.objects.get(
                anno=anno,
                activo=True
                # encargado=usuario_codigo <-- Eliminado
            )
        except ObjetivoAnual.DoesNotExist:
            return Response({"error": "No existe ese objetivo"}, status=status.HTTP_404_NOT_FOUND)

        serializer = ObjetivoAnualSerializer(objetivo, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

##=========##
## RESUMEN ##
##=========##
@api_view(["GET"])
def resumen_dashboard(request):

    usuario_codigo = request.user.usuario_usu
    nombre_corto = request.user.nomb_cort_usu

    anno = timezone.now().year
    mes = timezone.now().month

    # ===============================
    # OBJETIVO ANUAL
    # ===============================
    try:
        objetivo = ObjetivoAnual.objects.prefetch_related("areas").get(
            encargado=usuario_codigo,
            anno=anno,
            activo=True
        )
    except ObjetivoAnual.DoesNotExist:
        return Response({"message": "Sin objetivo configurado"})

    # ===============================
    # COTIZACIONES ANUALES
    # ===============================
    base = DashboardCotizacion.objects.filter(
        nombc=nombre_corto
    )

    base = aplicar_filtros(base, request)

    cotizaciones_anuales = base.filter(fecha__year=anno)
    cotizaciones_mes = base.filter(
        fecha__year=anno,
        fecha__month=mes
    )

    # ===============================
    # COTIZACIONES MENSUALES
    # ===============================
    cotizaciones_mes = DashboardCotizacion.objects.filter(
        nombc=nombre_corto,
        fecha__year=anno,
        fecha__month=mes
    ).values_list("num_reg", flat=True)

    # ===============================
    # EXPRESIÓN UTILIDAD
    # ===============================
    utilidad_expr = ExpressionWrapper(
        F("tou") * F("tde") * F("can"),
        output_field=DecimalField(max_digits=18, decimal_places=2)
    )

    # ===============================
    # ANUAL
    # ===============================
    hh_anual = CotiServicios.objects.filter(
        num_reg__in=cotizaciones_anuales,
        cog__regex=r"^\d{3}4",
        nig=2
    ).aggregate(total=Coalesce(Sum("toc"), Decimal("0")))["total"]

    utilidad_anual = CotiServicios.objects.filter(
        num_reg__in=cotizaciones_anuales,
        cog__regex=r"^\d{3}[46]"
    ).aggregate(total=Coalesce(Sum(utilidad_expr), Decimal("0")))["total"]

    logrado_anual = hh_anual + utilidad_anual

    # ===============================
    # MENSUAL
    # ===============================
    hh_mes = CotiServicios.objects.filter(
        num_reg__in=cotizaciones_mes,
        cog__regex=r"^\d{3}4",
        nig=2
    ).aggregate(total=Coalesce(Sum("toc"), Decimal("0")))["total"]

    utilidad_mes = CotiServicios.objects.filter(
        num_reg__in=cotizaciones_mes,
        cog__regex=r"^\d{3}[46]"
    ).aggregate(total=Coalesce(Sum(utilidad_expr), Decimal("0")))["total"]

    logrado_mes = hh_mes + utilidad_mes

    # ===============================
    # METAS
    # ===============================
    min_anual = sum(a.minimo for a in objetivo.areas.all())
    max_anual = sum(a.maximo for a in objetivo.areas.all())

    meta_mensual = min_anual / Decimal("12")

    # ===============================
    # SEMÁFOROS
    # ===============================
    def calcular_estado(valor, minimo, maximo):
        if valor < minimo:
            return "rojo"
        elif minimo <= valor < maximo:
            return "amarillo"
        return "verde"

    estado_anual = calcular_estado(logrado_anual, min_anual, max_anual)
    estado_mes = calcular_estado(logrado_mes, meta_mensual, max_anual / 12)

    # ===============================
    # RESPUESTA FINAL
    # ===============================
    return Response({
        "anual": {
            "logrado": logrado_anual,
            "min": min_anual,
            "max": max_anual,
            "estado": estado_anual
        },
        "mensual": {
            "logrado": logrado_mes,
            "meta": meta_mensual,
            "estado": estado_mes
        }
    })

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def logrado_dashboard(request):

    anno = int(request.GET.get("anno", datetime.now().year))
    mes_actual = str(datetime.now().month).zfill(2)  # 01, 02, etc

    print(f"DEBUG: Año = {anno}, Mes actual = {mes_actual}")

    # Cotizaciones aprobadas del año
    cotizaciones_anuales = DashboardCotizacion.objects.filter(
        anno=anno,
        envio="3",
    )
    print(f"DEBUG: Cotizaciones anuales encontradas = {cotizaciones_anuales.count()}")

    # Cotizaciones del mes actual
    cotizaciones_mensuales = cotizaciones_anuales.filter(
        mes=mes_actual
    )
    print(f"DEBUG: Cotizaciones mensuales encontradas = {cotizaciones_mensuales.count()}")

    utilidad_expr = ExpressionWrapper(
        F("tou") * F("can") * F("tde"),
        output_field=DecimalField(max_digits=18, decimal_places=2),
    )

    def calcular_logrado(cotizaciones):
        total_hh = Decimal("0.00")
        total_utilidad = Decimal("0.00")

        num_regs = [coti.num_reg for coti in cotizaciones]
        if not num_regs:
            return Decimal("0.00")

        # HH PROPIOS
        hh_agg = CotiServicios.objects.filter(
            num_reg__in=num_regs,
            cog__regex=r"^\d{3}4",
            nig=2,
        ).aggregate(
            total=Coalesce(Sum("toc"), Decimal("0.00"), output_field=DecimalField(max_digits=18, decimal_places=2))
        )
        total_hh = hh_agg["total"]

        # UTILIDAD
        utilidad_agg = CotiServicios.objects.filter(
            num_reg__in=num_regs,
        ).aggregate(
            total=Coalesce(Sum(utilidad_expr), Decimal("0.00"))
        )
        total_utilidad = utilidad_agg["total"]

        return total_hh + total_utilidad

    total_anual = calcular_logrado(cotizaciones_anuales)
    total_mensual = calcular_logrado(cotizaciones_mensuales)

    print(f"DEBUG: Total anual = {total_anual}, Total mensual = {total_mensual}")

    return JsonResponse({
        "anual": float(total_anual),
        "mensual": float(total_mensual),
    })

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def kpis_dashboard(request):
    from datetime import datetime
    from django.db.models import Sum, Count
    from django.db.models.functions import Coalesce, ExtractMonth
    from decimal import Decimal

    anno = int(request.GET.get("anno", datetime.now().year))
    mes_actual_num = datetime.now().month
    mes_actual_str = str(mes_actual_num).zfill(2)
    
    # Manejo de mes anterior para variaciones
    mes_anterior_num = 12 if mes_actual_num == 1 else mes_actual_num - 1
    anno_para_mes_anterior = anno - 1 if mes_actual_num == 1 else anno
    mes_anterior_str = str(mes_anterior_num).zfill(2)

    # ==========================================
    # 1. QUERIES BASE (Cotizaciones y Ventas)
    # ==========================================
    # Cotizaciones (Ofertas)
    qs_cot_anual = DashboardCotizacion.objects.filter(anno=anno)
    qs_cot_mes = qs_cot_anual.filter(mes=mes_actual_str)
    qs_cot_prev = DashboardCotizacion.objects.filter(anno=anno_para_mes_anterior, mes=mes_anterior_str)

    # Ventas Reales (Órdenes de Compra Adjudicadas oesta=1)
    qs_ventas_anual = vc_mov_orden.objects.filter(anno_a=str(anno), oesta=1)
    qs_ventas_mes = qs_ventas_anual.annotate(m=ExtractMonth("ofec")).filter(m=mes_actual_num)
    qs_ventas_prev = vc_mov_orden.objects.filter(anno_a=str(anno_para_mes_anterior), oesta=1)\
                        .annotate(m=ExtractMonth("ofec")).filter(m=mes_anterior_num)

    # ==========================================
    # 2. CÁLCULOS DE MÉTRICAS
    # ==========================================
    def get_monto(qs, field="tot_c"):
        return qs.aggregate(total=Coalesce(Sum(field), Decimal("0.00")))["total"]

    def calc_var(actual, anterior):
        if anterior and anterior != 0:
            return round(((float(actual) - float(anterior)) / float(anterior)) * 100, 1)
        return 0

    # --- KPI 1: Cantidad de Cotizaciones ---
    cant_mes = qs_cot_mes.count()
    cant_anual = qs_cot_anual.count()
    var_cant = calc_var(cant_mes, qs_cot_prev.count())

    # --- KPI 2: Monto Cotizado ---
    monto_cot_mes = get_monto(qs_cot_mes)
    monto_cot_anual = get_monto(qs_cot_anual)
    var_monto_cot = calc_var(monto_cot_mes, get_monto(qs_cot_prev))

    # --- KPI 3: Ventas Reales (OC) ---
    monto_v_mes = get_monto(qs_ventas_mes, "otot")
    monto_v_anual = get_monto(qs_ventas_anual, "otot")
    var_v = calc_var(monto_v_mes, get_monto(qs_ventas_prev, "otot"))

    # --- KPI 4: Efectividad (% Conversión de Monto) ---
    # Calculamos qué porcentaje del monto cotizado se convirtió en venta real
    def calc_efec(venta, coti):
        return round((float(venta) / float(coti) * 100), 1) if coti > 0 else 0

    efec_mes = calc_efec(monto_v_mes, monto_cot_mes)
    efec_anual = calc_efec(monto_v_anual, monto_cot_anual)

    # ==========================================
    # 3. RESPUESTA ESTRUCTURADA PARA EL FRONTEND
    # ==========================================
    return JsonResponse({
        # Cantidad de documentos
        "total_cotizaciones": {
            "anual": cant_anual,
            "variacion": var_cant
        },
        "cotizaciones_mes": cant_mes,

        # Monto ofertado
        "monto_total": {
            "anual": float(monto_cot_anual),
            "variacion": var_monto_cot
        },
        "monto_mes": float(monto_cot_mes),

        # Venta Real (OC)
        "ventas_reales_anual": float(monto_v_anual),
        "ventas_reales_mes": float(monto_v_mes),
        "ventas_variacion": var_v,

        # Ratios de eficiencia
        "porcentaje_aprobacion": efec_anual,
        "porcentaje_aprobacion_mes": efec_mes,
        
        # Extra (opcional por si lo usas en el footer)
        "ticket_promedio": float(monto_v_anual / qs_ventas_anual.count()) if qs_ventas_anual.count() > 0 else 0
    })

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def tendencias_dashboard(request):
    # 1. Obtener el año del request
    anno_buscado = str(request.GET.get("anno", datetime.now().year))

    # 2. DEFINIR LA BASE DE COTIZACIONES (Lo que se ofertó)
    base = DashboardCotizacion.objects.filter(num_reg__startswith=anno_buscado)
    base = aplicar_filtros(base, request)

    # --- LÓGICA DE VENTAS REALES (OC) ---
    # Filtramos por el campo anno_a y solo ADJUDICADAS (oesta=1)
    ventas_reales_qs = (
        vc_mov_orden.objects.filter(
            anno_a=anno_buscado, 
            oesta=1
        )
        .annotate(mes_num=ExtractMonth("ofec"))
        .values("mes_num")
        .annotate(total_oc=Coalesce(Sum("otot"), Decimal("0.00")))
    )
    
    ventas_reales_dict = {
        str(item["mes_num"]).zfill(2): item["total_oc"] 
        for item in ventas_reales_qs
    }

    # ============================
    # 1️⃣ VENTAS MENSUALES
    # ============================
    cotizaciones_raw = (
        base.values("mes")
        .annotate(
            total=Coalesce(Sum("tot_c"), Decimal("0.00")),
            cantidad=Count("num_reg")
        )
        .order_by("mes")
    )
    
    cot_mensuales_dict = {item["mes"]: item for item in cotizaciones_raw}
    
    ventas_mensuales = []
    for m in range(1, 13):
        mes_str = str(m).zfill(2)
        cot_data = cot_mensuales_dict.get(mes_str, {})
        
        ventas_mensuales.append({
            "mes": mes_str,
            "total": float(cot_data.get("total") or 0),
            "oc": float(ventas_reales_dict.get(mes_str) or 0),
            "cantidad": cot_data.get("cantidad") or 0
        })

    # ============================
    # 2️⃣ TOP VENTAS COMERCIAL
    # ============================
    dnis_permitidos = ['43662598', '20068421', '70942025']
    usuarios_qs = SegUsuario.objects.filter(dni__in=dnis_permitidos).values('dni', 'nomb_cort_usu', 'usuario_usu')
    mapa_usuarios = {u['dni']: (u['nomb_cort_usu'] or u['usuario_usu']).strip().upper() for u in usuarios_qs}

    cotizados_raw = (
        base.filter(envio=3, codic__in=dnis_permitidos)
        .values("codic")
        .annotate(
            monto_cotizado=Coalesce(Sum("tot_c"), Decimal("0.00")),
            cantidad_cot=Count("num_reg")
        )
    )
    dict_cotizados = {item['codic']: item for item in cotizados_raw}

    # Filtro de VENTAS por VENDEDOR (Agregando oesta=1)
    base_vendedores = base.filter(codic__in=dnis_permitidos)
    cotizaciones_ids = base_vendedores.values_list('numero', flat=True) 
    
    ventas_raw = (
        vc_mov_orden.objects.filter(
            cotin__in=cotizaciones_ids, 
            oesta=1  # <--- SOLO ADJUDICADAS
        )
        .exclude(otot__isnull=True)
        .values('cotin') 
        .annotate(total_venta_cotin=Sum('otot'))
    )
    dict_ventas_monto = {item['cotin']: item['total_venta_cotin'] for item in ventas_raw}

    ranking_comercial_unificado = []
    for codic, data_cot in dict_cotizados.items():
        nombre_vendedor = mapa_usuarios.get(codic, f"DNI: {codic}")
        ids_vendedor = base_vendedores.filter(codic=codic).values_list('numero', flat=True)
        
        venta_total = sum(float(dict_ventas_monto.get(cotin, 0) or 0) for cotin in ids_vendedor)
        monto_cotizado = float(data_cot['monto_cotizado'])
        cantidad = data_cot['cantidad_cot']

        ranking_comercial_unificado.append({
            "vendedor": nombre_vendedor,
            "monto": venta_total,
            "cotizado": monto_cotizado,
            "cantidad": cantidad,
            "ticket_promedio": venta_total / cantidad if cantidad > 0 else 0,
            "color": "#008B8B" 
        })

    ranking_comercial = sorted(ranking_comercial_unificado, key=lambda x: x['monto'], reverse=True)

    # ============================
    # 3️⃣ EMBUDO
    # ============================
    embudo = [
        {"etapa": "Cotizadas", "valor": base.count()},
        {"etapa": "Aprobadas", "valor": base.filter(envio=3).count()},
    ]

    # ============================
    # 4️⃣ DISTRIBUCIÓN POR ÁREA (Agregando oesta=1)
    # ============================
    AREA_MAP = {"1": "IND", "2": "MIN", "4": "OIL", "8": "SFY"}
    areas_cotizadas_raw = base.values("area_codigo").exclude(area_codigo="3").annotate(
        total_proyectos=Count("num_reg"),
        monto_cotizado=Coalesce(Sum("tot_c"), Decimal("0.00"))
    )

    areas_final = []
    for a in areas_cotizadas_raw:
        cod_area = str(a["area_codigo"])
        if cod_area not in AREA_MAP: continue
            
        ids_cotizaciones_area = base.filter(area_codigo=cod_area).values_list('numero', flat=True)
        
        # Filtro de Venta Real por Área con oesta=1
        venta_real_area = vc_mov_orden.objects.filter(
            cotin__in=ids_cotizaciones_area,
            oesta=1  # <--- SOLO ADJUDICADAS
        ).aggregate(total=Sum('otot'))['total'] or Decimal("0.00")

        areas_final.append({
            "area": AREA_MAP.get(cod_area),
            "total": a["total_proyectos"],
            "cotizado": float(a["monto_cotizado"]),
            "monto": float(venta_real_area)
        })

    areas_final = sorted(areas_final, key=lambda x: x['monto'], reverse=True)

    # ============================
    # 5️⃣ CLIENTES RECURRENTES
    # ============================
    agrupados_qs = (
        base.values("cliente_codigo")
        .annotate(
            total_cotizaciones=Count("num_reg"),
            monto_cotizado=Coalesce(Sum("tot_c"), Decimal("0.00"))
        )
        .order_by("-total_cotizaciones")[:10]
    )

    # PASO 2: Mapeo de nombres (Convertimos el entero a string para comparar)
    # Extraemos los códigos y evitamos errores si hay Nones
    codigos_top = [item["cliente_codigo"].strip() for item in agrupados_qs if item["cliente_codigo"]]
    
    clientes_db = vc_tab_clientes.objects.filter(codigo__in=codigos_top)
    
    # IMPORTANTE: Convertimos c.codigo a str() antes de hacer .strip()
    mapa_nombres = {
        str(c.codigo).strip(): c.nombre.strip() 
        for c in clientes_db
    }

    clientes_final = []
    for item in agrupados_qs:
        codigo_raw = item["cliente_codigo"] or ""
        codigo_limpio = codigo_raw.strip()
        
        # Buscamos en el mapa usando el string limpio
        nombre_real = mapa_nombres.get(codigo_limpio) or f"Cod: {codigo_limpio}"

        # PASO 3: Ventas Reales
        ids_cotizaciones_cliente = base.filter(cliente_codigo=codigo_raw).values_list('numero', flat=True)
        
        venta_real_cliente = vc_mov_orden.objects.filter(
            cotin__in=ids_cotizaciones_cliente,
            oesta=1
        ).aggregate(total=Sum('otot'))['total'] or Decimal("0.00")

        monto_c = float(item["monto_cotizado"])
        monto_v = float(venta_real_cliente)

        clientes_final.append({
            "codigo": codigo_limpio,
            "nombre": nombre_real,
            "cotizaciones": item["total_cotizaciones"],
            "monto_cotizado": monto_c,
            "monto_real": monto_v,
            "conversion": round((monto_v / monto_c * 100), 1) if monto_c > 0 else 0
        })

    clientes_recurrentes = sorted(clientes_final, key=lambda x: x['monto_real'], reverse=True)

    return Response({
        "ventas_mensuales": list(ventas_mensuales),
        "ranking_comercial": ranking_comercial,
        "embudo": embudo,
        "areas": areas_final,
        "clientes_recurrentes": clientes_recurrentes,
    })

def aplicar_filtros(base, request):

    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")
    area = request.GET.get("area")
    usuario = request.GET.get("usuario")
    tipo = request.GET.get("tipo")  # P, S, V

    # Rango de fechas
    if fecha_inicio and fecha_fin:
        base = base.filter(
            fecha__range=[fecha_inicio, fecha_fin]
        )

    # Área
    if area:
        base = base.filter(area_codigo=area)

    # Comercial (nombc guarda el nombre corto)
    if usuario:
        base = base.filter(nombc=usuario)

    # Tipo de cotización
    if tipo:
        base = base.filter(cotit=tipo)

    return base

##==========##
## ANALISIS ##
##==========##
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def cotizaciones_analisis_view(request):
    try:
        from django.db.models import Sum, Count, Avg
        from datetime import date

        dimension = request.data.get("dimension")
        metrica = request.data.get("metrica", "monto")
        filtros = request.data.get("filtros", {})

        qs = DashboardCotizacion.objects.all()

        # ==========================
        # Filtros globales
        # ==========================
        anno = filtros.get("anno", date.today().year)
        mes = filtros.get("mes")

        qs = qs.filter(fecha__year=anno)

        if mes:
            qs = qs.filter(fecha__month=mes)

        # Aquí puedes reutilizar tu lógica de filtros actual
        # cliente, estado, area, etc.

        # ==========================
        # Dimensión dinámica
        # ==========================
        DIMENSION_MAP = {
            "area": "area_codigo",
            "estado": "estado_nombre",
            "cliente": "cliente_nombre",
            "moneda": "tmone",
            "mes": "fecha__month",
            "vendedor": "nombt",
        }

        campo = DIMENSION_MAP.get(dimension)

        if not campo:
            return Response({"error": "Dimensión no válida"}, status=400)

        # ==========================
        # Métrica dinámica
        # ==========================
        if metrica == "monto":
            agg = Sum("tot_c")
        elif metrica == "cantidad":
            agg = Count("num_reg")
        elif metrica == "promedio":
            agg = Avg("tot_c")
        else:
            agg = Sum("tot_c")

        data = (
            qs.values(campo)
            .annotate(valor=agg)
            .order_by("-valor")
        )

        # ==========================
        # Formato frontend
        # ==========================
        resultado = []

        for r in data:
            resultado.append({
                dimension: r[campo],
                "valor": round(float(r["valor"] or 0), 2)
            })

        return Response(resultado)

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return Response({"error": str(e)}, status=500)

##================##
## NOTIFICACIONES ##
##================##
# OBTEER NOTIFACIONES
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def notificaciones_usuario(request):

    usuario = request.user

    notificaciones = Notificacion.objects.filter(
        usuario=usuario
    )[:10]

    serializer = NotificacionSerializer(notificaciones, many=True)

    return Response(serializer.data)

# MARCAR COMO LEIDO
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def marcar_notificacion(request, pk):

    try:
        notif = Notificacion.objects.get(pk=pk, usuario=request.user)
        notif.leido = True
        notif.save()
        return Response({"ok": True})
    except Notificacion.DoesNotExist:
        return Response({"error": "No encontrada"}, status=404)

# NO LEIDAS
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def notificaciones_no_leidas(request):

    total = Notificacion.objects.filter(
        usuario=request.user,
        leido=False
    ).count()

    return Response({"total": total})

# MAARCAR TODAS
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def marcar_todas_notificaciones(request):

    Notificacion.objects.filter(
        usuario=request.user,
        leido=False
    ).update(leido=True)

    return Response({"ok": True})

# GENERADOR
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def generar_alertas(request):

    alertas_sin_respuesta()

    # Aquí luego agregarás:
    # alertas_caida_conversion()
    # alertas_clientes_inactivos()
    # alertas_ticket_promedio()

    return Response({"ok": True})

# SIN RESPUESTA
def alertas_sin_respuesta():

    hoy = timezone.now().date()
    limite = hoy - timedelta(days=7)

    cotis = DashboardCotizacion.objects.filter(
        fecha__lte=limite,
        envio__isnull=True
    )

    for usuario in seg_usuario.objects.filter(activo=1):

        cantidad = cotis.filter(
            nombc=usuario.nomb_cort_usu
        ).count()

        if cantidad > 0:

            existe = Notificacion.objects.filter(
                usuario=usuario,
                titulo="Cotizaciones sin respuesta",
                leido=False
            ).exists()

            if not existe:
                Notificacion.objects.create(
                    usuario=usuario,
                    tipo="urgente",
                    titulo="Cotizaciones sin respuesta",
                    descripcion=f"Tienes {cantidad} cotizaciones con más de 7 días sin respuesta",
                    cantidad=cantidad
                )

#========================================================================================

##================##
## DATOS DE BD_VC ##
##================##
# vc_tab_areas
@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def lista_areas(request):
    if request.method == "GET":
        # Quitamos el filtro de activo para ver todo el catálogo
        areas = vc_tab_areas.objects.all().order_by("codigo")
        serializer = AreasSerializer(areas, many=True)
        return Response(serializer.data)
    
    elif request.method == "POST":
        serializer = AreasSerializer(data=request.data)
        if serializer.is_valid(): # Corregido: is_valid()
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# vc_tab_cargos
@api_view(["GET", "POST", "PUT"])
@permission_classes([IsAuthenticated])
def lista_cargos(request):
    # --- GET: Listar todos los cargos ---
    if request.method == "GET":
        cargos = vc_tab_cargos.objects.all().order_by("codigo")
        serializer = CargosSerializer(cargos, many=True)
        return Response(serializer.data)
    
    # --- POST: Crear un nuevo cargo ---
    elif request.method == "POST":
        serializer = CargosSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    # --- PUT: Actualizar (buscamos por el código enviado en el body) ---
    elif request.method == "PUT":
        codigo = request.data.get("codigo")
        try:
            cargo = vc_tab_cargos.objects.get(pk=codigo)
            # partial=True permite actualizar solo algunos campos si fuera necesario
            serializer = CargosSerializer(cargo, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except vc_tab_cargos.DoesNotExist:
            return Response({"error": "Cargo no encontrado"}, status=status.HTTP_404_NOT_FOUND)

# vc_tab_clientes
@api_view(["GET", "POST", "PUT", "DELETE"])
@permission_classes([IsAuthenticated])
def lista_clientes(request):
    
    # 1. GET:
    if request.method == "GET":
        clientes = vc_tab_clientes.objects.all() 
        serializer = ClientesSerializer(clientes, many=True)
        return Response(serializer.data)

    # 2. POST:
    elif request.method == "POST":
        serializer = ClientesSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({
                "message": "Empresa registrada correctamente",
                "data": serializer.data
            }, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    # 3. PUT:
    elif request.method == "PUT":
        codigo = request.data.get("codigo")
        try:
            cliente = vc_tab_clientes.objects.get(pk=codigo)
            serializer = ClientesSerializer(cliente, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({
                    "message": "Empresa actualizada correctamente",
                    "data": serializer.data
                }, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except vc_tab_clientes.DoesNotExist:
            return Response({"error": "Empresa no encontrada"}, status=status.HTTP_404_NOT_FOUND)

    # 4. DELETE:
    elif request.method == "DELETE":
        codigo = request.data.get("codigo") or request.query_params.get("codigo")
        
        if not codigo:
             return Response({"error": "Debe proporcionar el código"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            cliente = vc_tab_clientes.objects.get(pk=codigo)
            cliente.delete()
            return Response({"message": "Empresa eliminada correctamente"}, status=status.HTTP_200_OK)
        except vc_tab_clientes.DoesNotExist:
            return Response({"error": "Empresa no encontrada"}, status=status.HTTP_404_NOT_FOUND)
        except Exception:
            return Response({"error": "No se puede eliminar: el registro tiene datos asociados"}, status=status.HTTP_400_BAD_REQUEST)

# vc_tab_clientes_d
@api_view(["GET", "POST", "PUT", "DELETE"])
@permission_classes([IsAuthenticated])
def lista_representantes(request):
    
    # 1. GET: Listar todos
    if request.method == "GET":
        representantes = vc_tab_clientes_d.objects.all()
        serializer = RepresentantesSerializer(representantes, many=True)
        return Response(serializer.data)
    
    # 2. POST: Registro simple (Igual que Clientes)
    elif request.method == "POST":
        # Ya no calculamos el Max('codigo') aquí. 
        # Usamos directamente request.data que ya trae el código desde el Modal.
        serializer = RepresentantesSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({
                "message": "Representante registrado correctamente",
                "data": serializer.data
            }, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    # 3. PUT: Actualizar
    elif request.method == "PUT":
        codigo = request.data.get("codigo")
        try:
            representante = vc_tab_clientes_d.objects.get(pk=codigo)
            serializer = RepresentantesSerializer(representante, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({
                    "message": "Representante actualizado correctamente",
                    "data": serializer.data
                }, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except vc_tab_clientes_d.DoesNotExist:
            return Response({"error": "Representante no encontrado"}, status=status.HTTP_404_NOT_FOUND)

    # 4. DELETE: Eliminar
    elif request.method == "DELETE":
        codigo = request.data.get("codigo") or request.query_params.get("codigo")
        
        if not codigo:
            return Response({"error": "Debe proporcionar el código"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            representante = vc_tab_clientes_d.objects.get(pk=codigo)
            representante.delete()
            return Response({"message": "Representante eliminado correctamente"}, status=status.HTTP_200_OK)
        except vc_tab_clientes_d.DoesNotExist:
            return Response({"error": "Representante no encontrado"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"error": f"Error al eliminar: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

# vc_tab_estado
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def lista_estados(request):
    estados = vc_tab_estado.objects.filter(activo=True, cot=1).order_by("nombre")
    serializer = EstadoSerializer(estados, many=True)
    return Response(serializer.data)

#vc_tab_categorias
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def lista_categorias(request):
    categorias = vc_tab_categorias.objects.filter(activo="1").order_by("nombre")
    serializer = CategoriasSerializer(categorias, many=True)
    return Response(serializer.data)

# vc_tab_tgastos
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def lista_tgasto(request):
    tgasto = vc_tab_tgastos.objects.filter(activo="1").order_by("codigo")
    serializer = TGastosSerializer(tgasto, many=True)
    return Response(serializer.data)

# vc_tab_tgastos_d
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def lista_tgasto_d(request):
    tgasto_d = vc_tab_tgastos_d.objects.filter(activo="1").order_by("nombre")
    serializer = TGastosDSerializer(tgasto_d, many=True)
    return Response(serializer.data)

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def lista_rittal(request):
    search = request.GET.get("search")
    limit = int(request.GET.get("limit", 15))

    queryset = vc_tab_rittal.objects.filter(activo="1")

    if search:
        queryset = queryset.filter(
            Q(nombre__icontains=search) |
            Q(codigo__icontains=search)
        )

    queryset = queryset.order_by("nombre")[:limit]
    serializer = RittalSerializer(queryset, many=True)
    return Response(serializer.data)

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def lista_rockwell(request):
    search = (request.GET.get("search") or "").strip()
    limit = int(request.GET.get("limit", 15))

    queryset = vc_tab_rockwell.objects.filter(activo="1")

    if search:
        queryset = queryset.filter(
            Q(codigo__icontains=search) |
            Q(codigo2__icontains=search) |
            Q(descripcion__icontains=search) |
            Q(ds__icontains=search)
        )

    queryset = queryset.order_by("codigo")[:limit]

    serializer = RockwellSerializer(queryset, many=True)
    return Response(serializer.data)

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def lista_ceyesa(request):
    search = request.GET.get("search")
    limit = int(request.GET.get("limit", 15))

    queryset = vc_tab_ceyesa.objects.filter(activo="1")

    if search:
        queryset = queryset.filter(
            Q(descripcion__icontains=search) |
            Q(codigo__icontains=search)
        )

    queryset = queryset.order_by("codigo")[:limit]
    serializer = CeyesaSerializer(queryset, many=True)
    return Response(serializer.data)

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def lista_hoffman(request):
    search = request.GET.get("search")
    limit = int(request.GET.get("limit", 15))

    queryset = vc_tab_hoffman.objects.filter(activo="1")

    if search:
        queryset = queryset.filter(
            Q(nombre__icontains=search) |
            Q(codigo__icontains=search)
        )

    queryset = queryset.order_by("nombre")[:limit]
    serializer = HoffmanSerializer(queryset, many=True)
    return Response(serializer.data)

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def lista_alm_articulos(request):
    search = request.GET.get("search", "").strip()
    proveedor = request.GET.get("proveedor")  # OTROS | Schneider | LS Industrial Systems
    limit = int(request.GET.get("limit", 15))

    queryset = alm_articulos.objects.all()

    # =========================
    # FILTRO POR PROVEEDOR
    # =========================
    if proveedor == "OTROS":
        queryset = queryset.filter(proveedor__isnull=True)
    elif proveedor:
        queryset = queryset.filter(proveedor__iexact=proveedor)

    # =========================
    # BUSQUEDA
    # =========================
    if search:
        queryset = queryset.filter(
            Q(nombre__icontains=search) |
            Q(codigo__icontains=search)
        )

    queryset = queryset.order_by("nombre")[:limit]

    serializer = AlmArticulosSerializer(queryset, many=True)
    return Response(serializer.data)

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def lista_proveedores(request):
    proveedores = vc_tab_tproveedor.objects.filter(activo="1").order_by("nombre")
    serializer = ProveedoresSerializer(proveedores, many=True)
    return Response(serializer.data)

# seg_usuario
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def listar_usuario(request):
    """
    Retorna los datos del usuario autenticado actual.
    """
    try:
        usuario = seg_usuario.objects.get(usuario_usu=request.user.username)
        serializer = SegUsuarioSerializer(usuario)
        return Response(serializer.data)
    except seg_usuario.DoesNotExist:
        return Response({"detail": "Usuario no encontrado."}, status=status.HTTP_404_NOT_FOUND)

# vc_tab_notas
@api_view(["GET", "POST", "PUT"])
@permission_classes([IsAuthenticated])
def lista_notas(request):
    # --- GET: Listar todas las notas ---
    if request.method == "GET":
        try:
            notas = vc_tab_notas.objects.all().order_by("codigo")
            serializer = NotasSerializer(notas, many=True)
            return Response(serializer.data)
        except Exception as e:
            return Response({"error_db": str(e)}, status=500)
    
    # --- POST: Crear una nueva nota ---
    elif request.method == "POST":
        serializer = NotasSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    # --- PUT: Actualizar nota técnica ---
    elif request.method == "PUT":
        codigo = request.data.get("codigo")
        try:
            nota = vc_tab_notas.objects.get(pk=codigo)
            # partial=True permite actualizar nombre o estado por separado
            serializer = NotasSerializer(nota, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except vc_tab_notas.DoesNotExist:
            return Response(
                {"error": "Nota técnica no encontrada"}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
#========================================================================================

##==========##
## REPORTES ##
##==========##
@csrf_exempt
def reporte_cotizaciones_dashboard_html(request):
    # =========================
    # Filtros
    # =========================
    anno    = request.GET.get("anno")
    mes     = request.GET.get("mes")
    estado  = request.GET.get("estado")
    cliente = request.GET.get("cliente")
    area    = request.GET.get("area")
    campo   = request.GET.get("campo")   # filtro general
    valor   = request.GET.get("valor")   # valor buscado

    qs = DashboardCotizacion.objects.all()

    if anno:
        qs = qs.filter(anno=anno)
    if mes and mes != "%":
        qs = qs.filter(mes=mes)
    if estado and estado != "%":
        qs = qs.filter(estado_codigo=estado)
    if cliente and cliente != "%":
        qs = qs.filter(cliente_codigo=cliente)
    if area and area != "%":
        qs = qs.filter(area_codigo=area)

    # Filtro general dinámico
    if campo and valor:
        filtro_dinamico = {f"{campo}__icontains": valor}
        qs = qs.filter(**filtro_dinamico)

    # =========================
    # Agrupación por área
    # =========================
    data = (
        qs.values("area_codigo")
        .annotate(
            cantidad=Count("num_reg"),
            importe=Sum("tot_c")
        )
        .order_by("area_codigo")
    )

    AREA_MAP = {
        "1": "Industria",
        "2": "Minería",
        "3": "Mantenimiento",
        "4": "Petroquímica",
        "8": "Seguridad de Maquinaria",
    }

    resultados = []
    total_general = Decimal("0.00")

    for row in data:
        importe = row["importe"] or Decimal("0.00")
        total_general += importe
        resultados.append({
            "area": AREA_MAP.get(row["area_codigo"], "Sin Área"),
            "cantidad": row["cantidad"],
            "importe": round(importe, 2),
        })

    context = {
        "resultados": resultados,
        "total_general": round(total_general, 2),
        "filtros": {
            "anno": anno,
            "mes": mes,
            "estado": estado,
            "cliente": cliente,
            "area": area,
            "campo": campo,
            "valor": valor,
        }
    }

    return render(request, "reportes/reporte_cotizaciones_dashboard.html", context)

@csrf_exempt
def reporte_suministros_html(request, num_reg):

    suministros = (
        CotiSuministros.objects
        .filter(num_reg=num_reg)
        .order_by("cog", "nig", "num")
    )

    if not suministros.exists():
        return HttpResponse(
            "No existen suministros para esta cotización",
            status=404
        )

    grupos = OrderedDict()

    # =========================
    # UTIL: PARSE COG
    # =========================
    def parse_cog(cog):
        cog = str(cog).zfill(4)
        tipo_code = cog[2:]  # 01, 02

        return {
            "01": "EQUIPOS",
            "02": "MATERIALES",
        }.get(tipo_code, "OTROS")

    # =========================
    # AGRUPAR POR COG
    # =========================
    for row in suministros:

        # Crear grupo si no existe
        if row.cog not in grupos:
            grupos[row.cog] = {
                "tipo": parse_cog(row.cog),  # EQUIPOS / MATERIALES
                "titulo": "",                # nog (ej: M1, M2)
                "items": []
            }

        # CABECERA (nig = 0)
        if row.nig == 0:
            grupos[row.cog]["titulo"] = row.nog or ""

        # ITEMS (nig > 0)
        elif row.nig > 0:
            utilidad = (row.tou or Decimal("0")) * (row.can or Decimal("0"))

            grupos[row.cog]["items"].append({
                "item": len(grupos[row.cog]["items"]) + 1,
                "cod": row.cod,
                "des": row.des,
                "pro": row.pro,
                "can": round(row.can or 0, 2),
                "val": round(row.val or 0, 2),
                "tot": round(row.tot or 0, 2),
                "puc": round(row.puc or 0, 2),
                "toc": round(row.toc or 0, 2),
                "utilidad": round(utilidad, 2),
            })

    context = {
        "num_reg": num_reg,
        "grupos": grupos
    }

    return render(
        request,
        "reportes/reporte_suministros.html",
        context
    )

@csrf_exempt
def reporte_suministros_excel(request, num_reg):
    suministros = CotiSuministros.objects.filter(num_reg=num_reg).order_by("cog", "nig", "num")
    if not suministros.exists():
        return HttpResponse("No existen suministros para esta cotización", status=404)

    grupos = OrderedDict()

    # Función para parsear cog
    def parse_cog(cog):
        cog = str(cog).zfill(4)
        contador = cog[:2]
        tipo_code = cog[2:]
        tipo_map = {
            "01": "EQUIPOS",
            "02": "MATERIALES",
        }
        tipo = tipo_map.get(tipo_code, "OTROS")
        subtitulo = f"M{int(contador)}"
        return tipo, subtitulo

    # Agrupar por COG
    for row in suministros:
        if row.cog not in grupos:
            tipo, subtitulo = parse_cog(row.cog)
            grupos[row.cog] = {
                "tipo": tipo,
                "subtitulo": subtitulo,
                "titulo": "",
                "items": []
            }

        # Cabecera (nig = 0)
        if row.nig == 0:
            grupos[row.cog]["titulo"] = row.nog or ""

        # Items (nig > 0)
        elif row.nig > 0:
            utilidad = (row.tou or Decimal("0")) * (row.can or Decimal("0"))
            grupos[row.cog]["items"].append({
                "item": len(grupos[row.cog]["items"]) + 1,
                "cod": row.cod,
                "des": row.des,
                "pro": row.pro,
                "can": round(row.can or 0, 2),
                "val": round(row.val or 0, 2),
                "tot": round(row.tot or 0, 2),
                "puc": round(row.puc or 0, 2),
                "toc": round(row.toc or 0, 2),
                "utilidad": round(utilidad, 2),
            })

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Suministros_{num_reg}"

    # Escribir tabla similar al HTML
    row_index = 1
    for grupo in grupos.values():
        # Título de grupo
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=10)
        ws.cell(row=row_index, column=1, value=f"{grupo['tipo']} - {grupo['titulo']}")
        row_index += 1

        # Cabecera
        headers_1 = ["Ítems", "Código", "Descripción", "Proveedor", "Cant", "Cliente Final", "", "Precio Final", "", "Utilidad"]
        headers_2 = ["", "", "", "", "", "Prec. Unit.", "Total", "Prec. Unit.", "Total", ""]
        ws.append(headers_1)
        ws.append(headers_2)
        row_index += 2

        # Items
        for it in grupo["items"]:
            ws.append([
                it["item"],
                it["cod"],
                it["des"],
                it["pro"],
                it["can"],
                it["val"],
                it["tot"],
                it["puc"],
                it["toc"],
                it["utilidad"]
            ])
            row_index += 1

        row_index += 1  # espacio entre grupos

    # Respuesta Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Suministros_{num_reg}.xlsx'
    wb.save(response)
    return response

@csrf_exempt
def reporte_servicios_html(request, num_reg):
    servicios = (
        CotiServicios.objects
        .filter(num_reg=num_reg)
        .order_by("cog", "nig", "num")
    )

    if not servicios.exists():
        return HttpResponse(
            "No existen servicios para esta cotización",
            status=404
        )

    grupos = OrderedDict()

    # =========================
    # AGRUPAR POR GRUPO Y SUBGRUPO
    # =========================
    for row in servicios:
        cog_key = str(row.cog).zfill(5)  # ej: 01041

        # Grupo principal (nig=0)
        if row.nig == 0:
            grupos[cog_key] = {
                "titulo": row.nog or "",
                "subgrupos": OrderedDict(),
                "total_val": Decimal("0.00"),
                "total_tot": Decimal("0.00"),
                "total_puc": Decimal("0.00"),
                "total_toc": Decimal("0.00"),
                "total_utilidad": Decimal("0.00"),
            }

        # Subgrupo (nig=1)
        elif row.nig == 1:
            # 🔥 Obtener tipo desde los dígitos centrales del COG (**04*, **05*, **06*)
            try:
                cog_str = str(row.cog).zfill(5)
                tipo_code = cog_str[2:4]  # ej: 01041 → "04"
            except:
                tipo_code = ""

            tipo_map = {
                "04": "MANO DE OBRA",
                "05": "GASTOS DE SERVICIOS",
                "06": "OTROS",
            }

            tipo = tipo_map.get(tipo_code, "OTROS")

            ultimo_grupo_key = list(grupos.keys())[-1]
            grupos[ultimo_grupo_key]["subgrupos"][cog_key] = {
                "tipo": tipo,
                "items": [],
                "sub_total_val": Decimal("0.00"),
                "sub_total_tot": Decimal("0.00"),
                "sub_total_puc": Decimal("0.00"),
                "sub_total_toc": Decimal("0.00"),
                "sub_total_utilidad": Decimal("0.00"),
            }


        # Items (nig=2)
        elif row.nig == 2:
            utilidad = (row.tou or Decimal("0")) * (row.tde or Decimal("0"))
            ultimo_grupo_key = list(grupos.keys())[-1]
            ultimo_subgrupo_key = list(grupos[ultimo_grupo_key]["subgrupos"].keys())[-1]

            item_data = {
                "item": len(grupos[ultimo_grupo_key]["subgrupos"][ultimo_subgrupo_key]["items"]) + 1,
                "cod": row.cod,
                "des": row.des,
                "pro": row.pro,
                "can": round(row.can or 0, 2),
                "dias": round(row.tde or 0, 0),
                "val": round(row.val or 0, 2),
                "tot": round(row.tot or 0, 2),
                "puc": round(row.puc or 0, 2),
                "toc": round(row.toc or 0, 2),
                "utilidad": round(utilidad, 2),
            }

            # Guardar item
            grupos[ultimo_grupo_key]["subgrupos"][ultimo_subgrupo_key]["items"].append(item_data)

            # Actualizar totales subgrupo
            sg = grupos[ultimo_grupo_key]["subgrupos"][ultimo_subgrupo_key]
            sg["sub_total_val"] += Decimal(str(item_data["val"]))
            sg["sub_total_tot"] += Decimal(str(item_data["tot"]))
            sg["sub_total_puc"] += Decimal(str(item_data["puc"]))
            sg["sub_total_toc"] += Decimal(str(item_data["toc"]))
            sg["sub_total_utilidad"] += Decimal(str(item_data["utilidad"]))

            # Actualizar totales grupo
            g = grupos[ultimo_grupo_key]
            g["total_val"] += Decimal(str(item_data["val"]))
            g["total_tot"] += Decimal(str(item_data["tot"]))
            g["total_puc"] += Decimal(str(item_data["puc"]))
            g["total_toc"] += Decimal(str(item_data["toc"]))
            g["total_utilidad"] += Decimal(str(item_data["utilidad"]))


    context = {
        "num_reg": num_reg,
        "grupos": grupos
    }

    return render(
        request,
        "reportes/reporte_servicios.html",
        context
    )

from django.db import connection

def sp_select_tabla_call(num_reg, area, tipo):
    with connection.cursor() as cursor:
        cursor.execute("CALL sp_select_tabla(%s, %s, %s)", [num_reg, area, tipo])
        row = cursor.fetchone()
        if row and row[0]:
            try:
                costo_str, ganancia_str = row[0].split("-")
                return Decimal(costo_str), Decimal(ganancia_str)
            except Exception:
                # Si el formato no es el esperado
                return Decimal("0"), Decimal("0")
        # Si no hay fila o es None
        return Decimal("0"), Decimal("0")

def sumatoria_costos(queryset):

    total_expr = ExpressionWrapper(
        F("tot") * F("can"),
        output_field=DecimalField(max_digits=14, decimal_places=2),
    )

    agg = queryset.aggregate(
        costo=Coalesce(Sum("toc"), Decimal("0")),
        total=Coalesce(Sum(total_expr), Decimal("0")),
    )

    ganancia = agg["total"] - agg["costo"]

    return agg["costo"], ganancia

@csrf_exempt
def reporte_detallado_cotizacion(request, num_reg):
    num_reg = str(num_reg)

    # ==========================================================
    # EQUIPOS (SU0 - nig=0 - cog 01%)
    # ==========================================================
    # 🔹 TOTAL VENTA
    qs_total = CotiSuministros.objects.filter(
        num_reg=num_reg,
        cog__endswith="01",
        nig=0,
    )

    total_expr = ExpressionWrapper(
        F("tot") * F("can"),
        output_field=DecimalField(max_digits=18, decimal_places=2),
    )

    agg_total = qs_total.aggregate(
        total=Coalesce(Sum(total_expr), Decimal("0"))
    )

    total_equipos = agg_total["total"]

    # 🔹 COSTO
    qs_costo = CotiSuministros.objects.filter(
        num_reg=num_reg,
        cog__endswith="01",
        nig=1,
    )

    agg_costo = qs_costo.aggregate(
        costo=Coalesce(Sum("toc"), Decimal("0"))
    )

    costo_equipos = agg_costo["costo"]

    # 🔹 GANANCIA
    ganancia_equipos = total_equipos - costo_equipos

    # ==========================================================
    # MATERIALES (02%)
    # ==========================================================
    # 🔹 TOTAL VENTA
    qs_total = CotiSuministros.objects.filter(
        num_reg=num_reg,
        cog__endswith="02",
        nig=0,
    )

    total_expr = ExpressionWrapper(
        F("tot") * F("can"),
        output_field=DecimalField(max_digits=18, decimal_places=2),
    )

    agg_total = qs_total.aggregate(
        total=Coalesce(Sum(total_expr), Decimal("0"))
    )

    total_materiales = agg_total["total"]

    # 🔹 COSTO
    qs_costo = CotiSuministros.objects.filter(
        num_reg=num_reg,
        cog__endswith="02",
        nig=1,
    )

    agg_costo = qs_costo.aggregate(
        costo=Coalesce(Sum("toc"), Decimal("0"))
    )

    costo_materiales = agg_costo["costo"]

    # 🔹 GANANCIA
    ganancia_materiales = total_materiales - costo_materiales

    # =========
    # AREA
    # =========
    from django.shortcuts import get_object_or_404

    cotizacion = get_object_or_404(DashboardCotizacion, num_reg=num_reg)

    area_cotizacion = str(cotizacion.area_codigo)

    # ==========================================================
    # HH PROPIOS (MOP → cog ???4)
    # ==========================================================
    qs = CotiServicios.objects.filter(
        num_reg=num_reg,
        cog__regex=r"^\d{3}4",
        tpr=area_cotizacion,
        nig=2,
    )

    total_hh_propios = qs.aggregate(
        total=Coalesce(
            Sum("tot"),
            Decimal("0.00"),
            output_field=DecimalField(max_digits=18, decimal_places=2),
        )
    )["total"]

    costo_hh_propios = qs.aggregate(
        costo=Coalesce(
            Sum("toc"),
            Decimal("0.00"),
            output_field=DecimalField(max_digits=18, decimal_places=2),
        )
    )["costo"]

    ganancia_hh_propios = total_hh_propios - costo_hh_propios

    # ==========================================================
    # HH OTRAS AREAS (MOI → cog ???4, área distinta)
    # ==========================================================
    qs_otras = CotiServicios.objects.filter(
        num_reg=num_reg,
        cog__regex=r"^\d{3}4",
        nig=2
    ).exclude(tpr=area_cotizacion)

    total_hh_otras = qs_otras.aggregate(
        total=Coalesce(
            Sum("tot"), 
            Decimal("0.00"),
            output_field=DecimalField(max_digits=18, decimal_places=2)
        )
    )["total"]

    costo_hh_otras = qs_otras.aggregate(
        costo=Coalesce(
            Sum("toc"), 
            Decimal("0.00"),
            output_field=DecimalField(max_digits=18, decimal_places=2)
        )
    )["costo"]

    ganancia_hh_otras = total_hh_otras - costo_hh_otras

    # ==========================================================
    # COSTO SERVICIOS (05%)
    # ==========================================================
    qs_base = CotiServicios.objects.filter(
        num_reg=num_reg,
        mov="05",
    )

    total_servicios = qs_base.filter(nig=1).aggregate(
        total=Coalesce(
            Sum("tot"),
            Decimal("0.00"),
            output_field=DecimalField(max_digits=18, decimal_places=2),
        )
    )["total"]

    costo_servicios = qs_base.filter(nig=2).aggregate(
        costo=Coalesce(
            Sum("toc"),
            Decimal("0.00"),
            output_field=DecimalField(max_digits=18, decimal_places=2),
        )
    )["costo"]

    ganancia_servicios = total_servicios - costo_servicios

    # ==========================================================
    # GASTOS ENTREGA (00%)
    # ==========================================================
    qs_base = CotiServicios.objects.filter(
        num_reg=num_reg,
        mov="00",
    )

    total_gastos_entrega = qs_base.filter(nig=1).aggregate(
        total=Coalesce(
            Sum("tot"),
            Decimal("0.00"),
            output_field=DecimalField(max_digits=18, decimal_places=2),
        )
    )["total"]

    costo_gastos_entrega = qs_base.filter(nig=2).aggregate(
        costo=Coalesce(
            Sum("toc"),
            Decimal("0.00"),
            output_field=DecimalField(max_digits=18, decimal_places=2),
        )
    )["costo"]

    ganancia_gastos_entrega = total_gastos_entrega - costo_gastos_entrega

    # ==========================================================
    # IMPREVISTOS (06%)
    # ==========================================================
    qs_base = CotiServicios.objects.filter(
        num_reg=num_reg,
        mov="06",
    )

    total_imprevistos = qs_base.filter(nig=1).aggregate(
        total=Coalesce(
            Sum("tot"),
            Decimal("0.00"),
            output_field=DecimalField(max_digits=18, decimal_places=2),
        )
    )["total"]

    costo_imprevistos = qs_base.filter(nig=2).aggregate(
        costo=Coalesce(
            Sum("toc"),
            Decimal("0.00"),
            output_field=DecimalField(max_digits=18, decimal_places=2),
        )
    )["costo"]

    ganancia_imprevistos = total_imprevistos - costo_imprevistos

    # ==========================================================
    # DESCUENTO
    # ==========================================================
    cotizacion = DashboardCotizacion.objects.only("des_m").get(num_reg=num_reg)

    costo_descuento = cotizacion.des_m or Decimal("0.00")

    ganancia_descuento = Decimal("0.00")

    total_descuento = -costo_descuento

    print(f"DEBUG: Area Cotizacion es {area_cotizacion}")
    print(f"DEBUG: Registros terminados en 6: {CotiServicios.objects.filter(num_reg=num_reg, cog__regex=r'^\d{3}6').count()}")
    
    # ==========================================================
    # TABLA FINAL
    # ==========================================================
    datos = [
        {"concepto": "EQUIPOS", "costo": costo_equipos, "ganancia": ganancia_equipos, "total": costo_equipos + ganancia_equipos},
        {"concepto": "MATERIALES", "costo": costo_materiales, "ganancia": ganancia_materiales, "total": costo_materiales + ganancia_materiales},
        {"concepto": "HH PROPIOS", "costo": costo_hh_propios, "ganancia": ganancia_hh_propios, "total": costo_hh_propios + ganancia_hh_propios},
        {"concepto": "COSTO SERVICIOS", "costo": costo_servicios, "ganancia": ganancia_servicios, "total": costo_servicios + ganancia_servicios},
        {"concepto": "GASTOS ENTREGA", "costo": costo_gastos_entrega, "ganancia": ganancia_gastos_entrega, "total": costo_gastos_entrega + ganancia_gastos_entrega},
        {"concepto": "IMPREVISTOS", "costo": costo_imprevistos, "ganancia": ganancia_imprevistos, "total": costo_imprevistos + ganancia_imprevistos},
        {"concepto": "DESCUENTO", "costo": costo_descuento, "ganancia": ganancia_descuento, "total": total_descuento},
    ]

    total_final = sum(d["total"] for d in datos)

    context = {
        "num_reg": num_reg,
        "titulo": "RESUMEN DE COSTO UTILIDAD",
        "proyecto": num_reg,
        "datos": datos,
        "total_final": total_final,
    }

    return render(
        request,
        "reportes/reporte_detallado_cotizacion.html",
        context,
    )

@csrf_exempt
def reporte_detallado_excel(request, num_reg):
    num_reg = str(num_reg)

    # === CALCULOS (igual que tu versión) ===
    qs_total = CotiSuministros.objects.filter(num_reg=num_reg, cog__endswith="01", nig=0)
    total_equipos = qs_total.aggregate(total=Coalesce(Sum(F("tot") * F("can")), Decimal("0")))["total"]
    qs_costo = CotiSuministros.objects.filter(num_reg=num_reg, cog__endswith="01", nig=1)
    costo_equipos = qs_costo.aggregate(costo=Coalesce(Sum("toc"), Decimal("0")))["costo"]
    ganancia_equipos = total_equipos - costo_equipos

    qs_total = CotiSuministros.objects.filter(num_reg=num_reg, cog__endswith="02", nig=0)
    total_materiales = qs_total.aggregate(total=Coalesce(Sum(F("tot") * F("can")), Decimal("0")))["total"]
    qs_costo = CotiSuministros.objects.filter(num_reg=num_reg, cog__endswith="02", nig=1)
    costo_materiales = qs_costo.aggregate(costo=Coalesce(Sum("toc"), Decimal("0")))["costo"]
    ganancia_materiales = total_materiales - costo_materiales

    cotizacion = DashboardCotizacion.objects.only("des_m", "area_codigo").get(num_reg=num_reg)
    area_cotizacion = str(cotizacion.area_codigo)

    qs = CotiServicios.objects.filter(num_reg=num_reg, cog__regex=r"^\d{3}4", tpr=area_cotizacion, nig=2)
    total_hh_propios = qs.aggregate(total=Coalesce(Sum("tot"), Decimal("0")))["total"]
    costo_hh_propios = qs.aggregate(costo=Coalesce(Sum("toc"), Decimal("0")))["costo"]
    ganancia_hh_propios = total_hh_propios - costo_hh_propios

    qs_total_otras = CotiServicios.objects.filter(num_reg=num_reg, cog__regex=r"^\d{3}4", nig=0)
    total_hh_otras = qs_total_otras.aggregate(total=Coalesce(Sum(F("tot") * F("can")), Decimal("0")))["total"]
    qs_costo_otras = CotiServicios.objects.filter(num_reg=num_reg, cog__regex=r"^\d{3}4", nig=1)
    costo_hh_otras = qs_costo_otras.aggregate(costo=Coalesce(Sum("toc"), Decimal("0")))["costo"]
    ganancia_hh_otras = total_hh_otras - costo_hh_otras

    qs_base = CotiServicios.objects.filter(num_reg=num_reg, mov="05")
    total_servicios = qs_base.filter(nig=1).aggregate(total=Coalesce(Sum("tot"), Decimal("0")))["total"]
    costo_servicios = qs_base.filter(nig=2).aggregate(costo=Coalesce(Sum("toc"), Decimal("0")))["costo"]
    ganancia_servicios = total_servicios - costo_servicios

    qs_base = CotiServicios.objects.filter(num_reg=num_reg, mov="00")
    total_gastos_entrega = qs_base.filter(nig=1).aggregate(total=Coalesce(Sum("tot"), Decimal("0")))["total"]
    costo_gastos_entrega = qs_base.filter(nig=2).aggregate(costo=Coalesce(Sum("toc"), Decimal("0")))["costo"]
    ganancia_gastos_entrega = total_gastos_entrega - costo_gastos_entrega

    qs_base = CotiServicios.objects.filter(num_reg=num_reg, mov="06")
    total_imprevistos = qs_base.filter(nig=1).aggregate(total=Coalesce(Sum("tot"), Decimal("0")))["total"]
    costo_imprevistos = qs_base.filter(nig=2).aggregate(costo=Coalesce(Sum("toc"), Decimal("0")))["costo"]
    ganancia_imprevistos = total_imprevistos - costo_imprevistos

    costo_descuento = cotizacion.des_m or Decimal("0.00")
    ganancia_descuento = Decimal("0.00")
    total_descuento = -costo_descuento

    datos = [
        {"concepto": "EQUIPOS", "costo": costo_equipos, "ganancia": ganancia_equipos, "total": costo_equipos + ganancia_equipos},
        {"concepto": "MATERIALES", "costo": costo_materiales, "ganancia": ganancia_materiales, "total": costo_materiales + ganancia_materiales},
        {"concepto": "HH PROPIOS", "costo": costo_hh_propios, "ganancia": ganancia_hh_propios, "total": costo_hh_propios + ganancia_hh_propios},
        {"concepto": "HH OTRAS AREAS", "costo": costo_hh_otras, "ganancia": ganancia_hh_otras, "total": costo_hh_otras + ganancia_hh_otras},
        {"concepto": "COSTO SERVICIOS", "costo": costo_servicios, "ganancia": ganancia_servicios, "total": costo_servicios + ganancia_servicios},
        {"concepto": "GASTOS ENTREGA", "costo": costo_gastos_entrega, "ganancia": ganancia_gastos_entrega, "total": costo_gastos_entrega + ganancia_gastos_entrega},
        {"concepto": "IMPREVISTOS", "costo": costo_imprevistos, "ganancia": ganancia_imprevistos, "total": costo_imprevistos + ganancia_imprevistos},
        {"concepto": "DESCUENTO", "costo": costo_descuento, "ganancia": ganancia_descuento, "total": total_descuento},
    ]

    # === GENERAR EXCEL CON DISEÑO ===
    wb = Workbook()
    ws = wb.active
    ws.title = f"Detalle_{num_reg}"

    # Cabecera
    ws.append(["Concepto", "Costo", "Ganancia", "Total"])
    header_fill = PatternFill("solid", fgColor="BDD7EE")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in ws[1]:
        cell.font = Font(bold=True, color="000000")
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill
        cell.border = thin_border

    # Filas de datos
    for d in datos:
        ws.append([d["concepto"], float(d["costo"]), float(d["ganancia"]), float(d["total"])])

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=4):
        row[0].alignment = Alignment(horizontal="left")
        for cell in row[1:]:
            cell.alignment = Alignment(horizontal="right")
        for cell in row:
            cell.border = thin_border

    # Autoajustar ancho de columnas
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 5

    # Respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Detalle_{num_reg}.xlsx'
    wb.save(response)
    return response

@csrf_exempt
def reporte_resumen_cotizacion(request, num_reg):
    num_reg = str(num_reg)

    # --------------------------
    # REUTILIZAR CÁLCULOS DETALLE
    # --------------------------
    # Equipos
    qs_total = CotiSuministros.objects.filter(num_reg=num_reg, cog__endswith="01", nig=0)
    total_equipos = qs_total.aggregate(total=Coalesce(Sum(F("tot")*F("can")), Decimal("0")))["total"]
    qs_costo = CotiSuministros.objects.filter(num_reg=num_reg, cog__endswith="01", nig=1)
    costo_equipos = qs_costo.aggregate(costo=Coalesce(Sum("toc"), Decimal("0")))["costo"]
    ganancia_equipos = total_equipos - costo_equipos

    # Materiales
    qs_total = CotiSuministros.objects.filter(num_reg=num_reg, cog__endswith="02", nig=0)
    total_materiales = qs_total.aggregate(total=Coalesce(Sum(F("tot")*F("can")), Decimal("0")))["total"]
    qs_costo = CotiSuministros.objects.filter(num_reg=num_reg, cog__endswith="02", nig=1)
    costo_materiales = qs_costo.aggregate(costo=Coalesce(Sum("toc"), Decimal("0")))["costo"]
    ganancia_materiales = total_materiales - costo_materiales

    # Área
    cotizacion = get_object_or_404(DashboardCotizacion, num_reg=num_reg)
    area_cotizacion = str(cotizacion.area_codigo)

    # HH propios
    qs_hh_propios = CotiServicios.objects.filter(num_reg=num_reg, cog__regex=r"^\d{3}4", tpr=area_cotizacion, nig=2)
    total_hh_propios = qs_hh_propios.aggregate(total=Coalesce(Sum("tot"), Decimal("0")))["total"]
    costo_hh_propios = qs_hh_propios.aggregate(costo=Coalesce(Sum("toc"), Decimal("0")))["costo"]
    ganancia_hh_propios = total_hh_propios - costo_hh_propios

    # HH otras áreas
    qs_hh_otras = CotiServicios.objects.filter(num_reg=num_reg, cog__regex=r"^\d{3}4", nig=1)
    costo_hh_otras = qs_hh_otras.aggregate(costo=Coalesce(Sum("toc"), Decimal("0")))["costo"]
    ganancia_hh_otras = CotiServicios.objects.filter(num_reg=num_reg, cog__regex=r"^\d{3}4", nig=0).aggregate(total=Coalesce(Sum(F("tot")*F("can")), Decimal("0")))["total"] - costo_hh_otras

    # Costo servicios
    qs_servicios = CotiServicios.objects.filter(num_reg=num_reg, mov="05")
    costo_servicios = qs_servicios.filter(nig=2).aggregate(costo=Coalesce(Sum("toc"), Decimal("0")))["costo"]
    ganancia_servicios = qs_servicios.filter(nig=1).aggregate(total=Coalesce(Sum("tot"), Decimal("0")))["total"] - costo_servicios

    # Gastos entrega
    qs_gastos_entrega = CotiServicios.objects.filter(num_reg=num_reg, mov="00", nig=2)
    costo_gastos_entrega = qs_gastos_entrega.aggregate(costo=Coalesce(Sum("toc"), Decimal("0")))["costo"]
    ganancia_gastos_entrega = qs_gastos_entrega.aggregate(total=Coalesce(Sum("tot"), Decimal("0")))["total"] - costo_gastos_entrega

    # Imprevistos
    qs_imprevistos = CotiServicios.objects.filter(num_reg=num_reg, mov="06", nig=1)
    costo_imprevistos = qs_imprevistos.aggregate(costo=Coalesce(Sum("toc"), Decimal("0")))["costo"]
    ganancia_imprevistos = qs_imprevistos.aggregate(total=Coalesce(Sum("tot"), Decimal("0")))["total"] - costo_imprevistos

    # --------------------------
    # RESUMEN
    # --------------------------
    resumen = [
        {"concepto": "GASTOS", "importe": costo_equipos + costo_servicios},
        {"concepto": "GANANCIAS", "importe": ganancia_equipos + ganancia_materiales + ganancia_hh_propios +
                                         ganancia_hh_otras + ganancia_imprevistos + ganancia_gastos_entrega + ganancia_servicios},
        {"concepto": "HH PROPIOS", "importe": costo_hh_propios},
        {"concepto": "IMPREVISTOS", "importe": costo_imprevistos},
        {"concepto": "GASTOS ENTREGA", "importe": costo_gastos_entrega},
    ]

    total_final = sum(d["importe"] for d in resumen)

    context = {
        "num_reg": num_reg,
        "titulo": "RESUMEN DE COSTO UTILIDAD",
        "proyecto": num_reg,
        "resumen": resumen,
        "total_final": total_final,
    }

    return render(
        request,
        "reportes/reporte_resumen_cotizacion.html",  # Plantilla específica para este resumen
        context,
    )

@csrf_exempt
def reporte_venta_total_html(request, num_reg):
    num_reg = str(num_reg)

    suministros = (
        CotiSuministros.objects
        .filter(num_reg=num_reg)
        .order_by("cog", "nig", "num")
    )

    if not suministros.exists():
        return HttpResponse("No existen suministros para esta cotización", status=404)

    grupos = OrderedDict()

    for row in suministros:
        if row.cog not in grupos:
            grupos[row.cog] = {
                "titulo_grupo": "", 
                "total_venta_grupo": Decimal("0.00"),
                "envio_grupo": Decimal("0.00"),
                "items": []
            }

        # CABECERA DEL GRUPO (nig = 0)
        if row.nig == 0:
            grupos[row.cog]["titulo_grupo"] = row.nog or "SIN TITULO"
            # Usamos env_tot que es el campo del modelo para el total del grupo
            grupos[row.cog]["envio_grupo"] = row.env_tot or Decimal("0.00")
            grupos[row.cog]["total_venta_grupo"] = row.tot or Decimal("0.00")

        # ITEMS DETALLE (nig > 0)
        else:
            subtotal_venta = row.tot or Decimal("0.00")
            subtotal_costo = row.toc or Decimal("0.00")
            
            grupos[row.cog]["items"].append({
                "cod": row.cod,
                "des": row.des,
                "can": row.can or Decimal("0"),
                "puc": row.puc or Decimal("0.00"),     # Costo Unitario Base
                "env_u": row.env_par or Decimal("0.00"), # Costo Envío por Item
                "cce": row.cost_c_env or Decimal("0.00"), # Costo Con Envío (ya calculado)
                "util_porc": row.cau or Decimal("0.00"), # % Utilidad (cau)
                "val": row.val or Decimal("0.00"),     # Precio Venta Unitario
                "tot": subtotal_venta,                  # Venta Total
                "util_money": subtotal_venta - subtotal_costo
            })

    context = {
        "num_reg": num_reg,
        "titulo": "REPORTE DETALLADO DE SUMINISTROS (VENTA TOTAL)",
        "grupos": grupos,
        "fecha": datetime.now()
    }

    return render(request, "reportes/reporte_venta_total.html", context)

@csrf_exempt
def reporte_venta_parcial_html(request, num_reg):
    num_reg = str(num_reg)
    
    # Traemos los datos ordenados por grupo y número de ítem
    suministros = (
        CotiSuministros.objects
        .filter(num_reg=num_reg)
        .order_by("cog", "nig", "num")
    )

    if not suministros.exists():
        return HttpResponse("No existen suministros para este reporte", status=404)

    grupos = OrderedDict()

    for row in suministros:
        if row.cog not in grupos:
            grupos[row.cog] = {
                "titulo_grupo": "", 
                "envio_grupo": Decimal("0.00"),
                "total_venta_grupo": Decimal("0.00"),
                "items": []
            }

        # CABECERA DEL GRUPO (nig = 0)
        if row.nig == 0:
            grupos[row.cog]["titulo_grupo"] = row.nog or "SIN TITULO"
            # env_tot representa el envío acumulado de este grupo
            grupos[row.cog]["envio_grupo"] = row.env_tot or Decimal("0.00")
            grupos[row.cog]["total_venta_grupo"] = row.tot or Decimal("0.00")

        # ITEMS DEL DETALLE (nig > 0)
        else:
            grupos[row.cog]["items"].append({
                "cod": row.cod,
                "des": row.des,
                "can": row.can or Decimal("0"),
                "puc": row.puc or Decimal("0.00"),     # Costo Unitario Base
                "toc": row.toc or Decimal("0.00"),     # Costo Total Base
                "env_u": row.cost_env or Decimal("0.00"), # Costo Envío (según modal)
                "cce": row.cost_c_env or Decimal("0.00"), # Costo con Envío
                "util_porc": row.cau or Decimal("0.00"),  # % Utilidad
                "util_money": row.tou or Decimal("0.00"), # Utilidad (monto unitario)
                "val": row.val or Decimal("0.00"),        # Venta Precio Unitario
                "tot": row.tot or Decimal("0.00")         # Venta Total
            })

    context = {
        "num_reg": num_reg,
        "titulo": "REPORTE DE SUMINISTROS (VENTA PARCIAL)",
        "grupos": grupos,
        "fecha": datetime.now()
    }
    
    return render(request, "reportes/reporte_venta_parcial.html", context)

#========================================================================================

##============##
## AÑO ACTUAL ##
##============##
@api_view(["GET"])
def anno_actual(request):
    try:
        cia = cont_cias.objects.get(cod="001")
        return Response({"anno": cia.anno})
    except cont_cias.DoesNotExist:
        return Response({"anno": timezone.now().year})

#========================================================================================

##===============##
## FUNCION TEXTO ##
##===============##
from bs4 import BeautifulSoup

def html_to_text(html):
    if not html:
        return ""

    soup = BeautifulSoup(html, "html.parser")

    # Reemplaza <li> por "- texto"
    for li in soup.find_all("li"):
        li.insert_before("- ")
        li.append("\n")

    # Convierte <p> en saltos de línea
    for p in soup.find_all("p"):
        p.append("\n")

    text = soup.get_text()
    return text.strip()